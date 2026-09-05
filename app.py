# app.py â€” Entry point Flask, semua route
# Komentar dalam Bahasa Indonesia sesuai konvensi (context.md bagian 12)

from flask import (Flask, request, jsonify, render_template,
                   redirect, url_for, session, flash)
from flask_cors import CORS
from flask_socketio import SocketIO, emit
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo
from uuid import uuid4
from config import APP_TIMEZONE

# Timezone WIB (UTC+7) â€” digunakan di semua fungsi waktu
WIB = ZoneInfo(APP_TIMEZONE)

def now_wib():
    """Dapatkan waktu sekarang dalam timezone WIB."""
    return datetime.now(WIB)
import os
import base64
import binascii
import json
import re
import threading
import time
import numpy as np
import cv2
import database as db
from config import (FLASK_HOST, FLASK_PORT, FLASK_DEBUG, FLASK_SECRET_KEY,
                    SNAPSHOT_PATH, TOLERANSI_MENIT, DATASET_PATH, FOTO_PER_USER,
                    CONFIDENCE_THRESHOLD, ANTI_SPOOFING_THRESHOLD,
                    ESP32_ENABLED, ESP32_IP, ESP32_PORT, ESP32_TIMEOUT,
                    MODEL_PATH, RECOGNITION_REQUIRED_FRAMES,
                    FACE_MATCH_THRESHOLD, ENROLLMENT_STABLE_FRAMES,
                    ENROLLMENT_STATE_TTL_SECONDS, FACE_GALLERY_META_PATH)

# Folder runtime harus tersedia juga ketika app dimuat melalui Gunicorn/WSGI.
os.makedirs(SNAPSHOT_PATH, exist_ok=True)
os.makedirs(DATASET_PATH, exist_ok=True)
model_dir = os.path.dirname(MODEL_PATH)
if model_dir:
    os.makedirs(model_dir, exist_ok=True)

# â”€â”€ Inisialisasi Flask + SocketIO â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
app = Flask(__name__)
app.secret_key = FLASK_SECRET_KEY
app.config['TEMPLATES_AUTO_RELOAD'] = True
CORS(app)
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

# Status kamera global
_camera_states = {}
_socket_tracker_keys = {}

# Tracker dipisahkan per client agar kamera tidak saling menimpa status.
_consecutive_trackers = {}
_completed_trackers = {}
_consecutive_lock = threading.Lock()
_training_lock = threading.Lock()
_gallery_build_state_lock = threading.RLock()
_gallery_build_state = {
    'build_id': None,
    'state': 'idle',
    'last_error': None,
    'started_at': None,
    'finished_at': None,
    'updated_at': None,
    'requested_user_id': None,
}
_enrollment_lock = threading.Lock()
_enrollment_states = {}
_enrollment_upload_locks = {}
_background_lock = threading.Lock()
_background_started = False


def _set_gallery_build_state(**changes):
    """Update the single-process gallery build lifecycle for UI polling."""
    with _gallery_build_state_lock:
        _gallery_build_state.update(changes)
        _gallery_build_state['updated_at'] = now_wib().isoformat()
        return dict(_gallery_build_state)


def _get_gallery_build_state():
    with _gallery_build_state_lock:
        return dict(_gallery_build_state)


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# DECORATOR: login_required
# Semua route kecuali /login dan /register wajib pakai ini
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'admin_id' not in session:
            if request.path.startswith('/api/'):
                return jsonify({
                    'status': 'error', 'data': None,
                    'pesan': 'Phiên đăng nhập đã hết hạn. Vui lòng đăng nhập lại.'
                }), 401
            flash('Vui lÃ²ng Ä‘Äƒng nháº­p trÆ°á»›c.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# AUTH: Login, Register, Logout
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Halaman login admin."""
    # Jika sudah login, langsung ke dashboard
    if 'admin_id' in session:
        return redirect(url_for('dashboard'))

    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        if not username or not password:
            error = 'Vui lÃ²ng nháº­p tÃªn Ä‘Äƒng nháº­p vÃ  máº­t kháº©u.'
        else:
            admin = db.get_admin_by_username(username)
            if admin and check_password_hash(admin['password_hash'], password):
                # Login berhasil â€” simpan ke session
                session['admin_id'] = admin['id']
                session['username'] = admin['username']
                flash('ÄÄƒng nháº­p thÃ nh cÃ´ng!', 'success')
                return redirect(url_for('dashboard'))
            else:
                error = 'TÃªn Ä‘Äƒng nháº­p hoáº·c máº­t kháº©u khÃ´ng Ä‘Ãºng.'

    # Tampilkan link register hanya jika belum ada admin sama sekali
    show_register = (db.hitung_admin() == 0)

    return render_template('login.html', error=error, show_register=show_register)


@app.route('/register', methods=['GET', 'POST'])
def register():
    """Halaman register admin pertama.
    Hanya bisa diakses jika tabel admin masih kosong (0 record).
    Jika sudah ada admin, redirect ke /login.
    """
    # Cek apakah sudah ada admin
    if db.hitung_admin() > 0:
        flash('Quáº£n trá»‹ viÃªn Ä‘Ã£ Ä‘Æ°á»£c Ä‘Äƒng kÃ½. Vui lÃ²ng Ä‘Äƒng nháº­p.', 'error')
        return redirect(url_for('login'))

    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        confirm  = request.form.get('confirm_password', '')

        # Validasi input
        if not username or not password:
            error = 'Vui lÃ²ng nháº­p tÃªn Ä‘Äƒng nháº­p vÃ  máº­t kháº©u.'
        elif len(password) < 8:
            error = 'Máº­t kháº©u pháº£i cÃ³ Ã­t nháº¥t 8 kÃ½ tá»±.'
        elif password != confirm:
            error = 'Máº­t kháº©u xÃ¡c nháº­n khÃ´ng khá»›p.'
        else:
            # Hash password dan simpan
            hashed = generate_password_hash(password)
            admin_id = db.tambah_admin(username, hashed)
            if admin_id:
                # Langsung login setelah register
                session['admin_id'] = admin_id
                session['username'] = username
                flash('Táº¡o tÃ i khoáº£n quáº£n trá»‹ viÃªn thÃ nh cÃ´ng!', 'success')
                return redirect(url_for('dashboard'))
            else:
                error = 'KhÃ´ng thá»ƒ táº¡o tÃ i khoáº£n. TÃªn Ä‘Äƒng nháº­p cÃ³ thá»ƒ Ä‘Ã£ Ä‘Æ°á»£c sá»­ dá»¥ng.'

    return render_template('register_admin.html', error=error)


@app.route('/logout')
def logout():
    """Logout admin â€” hapus session."""
    session.clear()
    flash('Báº¡n Ä‘Ã£ Ä‘Äƒng xuáº¥t.', 'success')
    return redirect(url_for('login'))


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# DASHBOARD
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@app.route('/')
@login_required
def dashboard():
    """Halaman utama dashboard."""
    tanggal_hari_ini = now_wib().date()
    statistik = db.get_statistik_dashboard(tanggal_hari_ini)
    absensi = db.get_absensi_hari_ini(tanggal_hari_ini)
    return render_template('dashboard.html',
                           active_page='dashboard',
                           statistik=statistik,
                           absensi_hari_ini=absensi,
                           conf_threshold=FACE_MATCH_THRESHOLD,
                           spoof_threshold=ANTI_SPOOFING_THRESHOLD)


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# MANAJEMEN KELAS
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@app.route('/kelas')
@login_required
def kelas_index():
    """Daftar semua kelas."""
    daftar = db.get_semua_kelas()
    # Tambahkan jumlah mahasiswa per kelas
    for k in daftar:
        k['jumlah_mhs'] = db.hitung_mahasiswa_per_kelas(k['id'])
    return render_template('kelas/index.html',
                           active_page='kelas', daftar_kelas=daftar)


@app.route('/kelas/tambah', methods=['GET', 'POST'])
@login_required
def kelas_tambah():
    """Tambah kelas baru."""
    error = None
    if request.method == 'POST':
        nama = request.form.get('nama_kelas', '').strip()
        angkatan = request.form.get('angkatan', '').strip()
        if not nama or not angkatan:
            error = 'Vui lÃ²ng nháº­p tÃªn lá»›p vÃ  khÃ³a há»c.'
        else:
            hasil = db.tambah_kelas(nama, angkatan, session.get('admin_id'))
            if hasil:
                flash('ThÃªm lá»›p thÃ nh cÃ´ng!', 'success')
                return redirect(url_for('kelas_index'))
            error = 'KhÃ´ng thá»ƒ thÃªm lá»›p.'
    return render_template('kelas/form.html',
                           active_page='kelas', kelas=None, error=error)


@app.route('/kelas/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def kelas_edit(id):
    """Edit kelas."""
    kelas = db.get_kelas_by_id(id)
    if not kelas:
        flash('KhÃ´ng tÃ¬m tháº¥y lá»›p.', 'error')
        return redirect(url_for('kelas_index'))
    error = None
    if request.method == 'POST':
        nama = request.form.get('nama_kelas', '').strip()
        angkatan = request.form.get('angkatan', '').strip()
        if not nama or not angkatan:
            error = 'Vui lÃ²ng nháº­p tÃªn lá»›p vÃ  khÃ³a há»c.'
        elif db.update_kelas(id, nama, angkatan):
            flash('Cáº­p nháº­t lá»›p thÃ nh cÃ´ng!', 'success')
            return redirect(url_for('kelas_index'))
        else:
            error = 'KhÃ´ng thá»ƒ cáº­p nháº­t lá»›p.'
    return render_template('kelas/form.html',
                           active_page='kelas', kelas=kelas, error=error)


@app.route('/kelas/hapus/<int:id>', methods=['POST'])
@login_required
def kelas_hapus(id):
    """Hapus kelas (CASCADE ke MK dan jadwal)."""
    if db.hapus_kelas(id):
        flash('XÃ³a lá»›p thÃ nh cÃ´ng.', 'success')
    else:
        flash('KhÃ´ng thá»ƒ xÃ³a lá»›p. CÃ³ thá»ƒ lá»›p váº«n cÃ²n sinh viÃªn liÃªn quan.', 'error')
    return redirect(url_for('kelas_index'))


@app.route('/kelas/<int:id>/matakuliah')
@login_required
def kelas_detail(id):
    """Detail kelas + daftar matakuliah."""
    kelas = db.get_kelas_by_id(id)
    if not kelas:
        flash('KhÃ´ng tÃ¬m tháº¥y lá»›p.', 'error')
        return redirect(url_for('kelas_index'))
    matakuliah = db.get_matakuliah_by_kelas(id)
    jumlah_mhs = db.hitung_mahasiswa_per_kelas(id)
    return render_template('kelas/detail.html',
                           active_page='kelas', kelas=kelas,
                           matakuliah=matakuliah, jumlah_mhs=jumlah_mhs)


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# MANAJEMEN MATAKULIAH
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@app.route('/matakuliah/tambah', methods=['GET', 'POST'])
@login_required
def matakuliah_tambah():
    """Tambah matakuliah baru."""
    kelas_id_param = request.args.get('kelas_id', type=int)
    kelas_asal = db.get_kelas_by_id(kelas_id_param) if kelas_id_param else None
    error = None
    if request.method == 'POST':
        nama = request.form.get('nama_mk', '').strip()
        kode = request.form.get('kode_mk', '').strip()
        kid  = request.form.get('kelas_id', type=int)
        sks  = request.form.get('sks', 2, type=int)
        if not nama or not kode or not kid:
            error = 'Vui lÃ²ng nháº­p Ä‘áº§y Ä‘á»§ thÃ´ng tin.'
        else:
            hasil = db.tambah_matakuliah(nama, kode, kid, sks)
            if hasil:
                flash('ThÃªm mÃ´n há»c thÃ nh cÃ´ng!', 'success')
                return redirect(url_for('kelas_detail', id=kid))
            error = 'KhÃ´ng thá»ƒ thÃªm mÃ´n há»c. MÃ£ mÃ´n há»c cÃ³ thá»ƒ Ä‘Ã£ Ä‘Æ°á»£c sá»­ dá»¥ng.'
            error = 'KhÃ´ng thá»ƒ thÃªm mÃ´n há» c. MÃ£ mÃ´n há» c cÃ³ thá»ƒ Ä‘Ã£ Ä‘Æ°á»£c sá»­ dá»¥ng.'
    return render_template('matakuliah/form.html', active_page='kelas',
                           mk=None, kelas_asal=kelas_asal,
                           daftar_kelas=db.get_semua_kelas(), error=error)


@app.route('/matakuliah/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def matakuliah_edit(id):
    """Edit matakuliah."""
    mk = db.get_matakuliah_by_id(id)
    if not mk:
        flash('KhÃ´ng tÃ¬m tháº¥y mÃ´n há» c.', 'error')
        return redirect(url_for('kelas_index'))
    kelas_asal = db.get_kelas_by_id(mk['kelas_id'])
    error = None
    if request.method == 'POST':
        nama = request.form.get('nama_mk', '').strip()
        kode = request.form.get('kode_mk', '').strip()
        kid  = request.form.get('kelas_id', type=int)
        sks  = request.form.get('sks', 2, type=int)
        if db.update_matakuliah(id, nama, kode, kid, sks):
            flash('Cáº­p nháº­t mÃ´n há»c thÃ nh cÃ´ng!', 'success')
            return redirect(url_for('kelas_detail', id=kid))
        error = 'KhÃ´ng thá»ƒ cáº­p nháº­t mÃ´n há»c. MÃ£ mÃ´n há»c cÃ³ thá»ƒ Ä‘Ã£ Ä‘Æ°á»£c sá»­ dá»¥ng.'
    return render_template('matakuliah/form.html', active_page='kelas',
                           mk=mk, kelas_asal=kelas_asal,
                           daftar_kelas=db.get_semua_kelas(), error=error)


@app.route('/matakuliah/hapus/<int:id>', methods=['POST'])
@login_required
def matakuliah_hapus(id):
    """Hapus matakuliah (CASCADE ke jadwal)."""
    mk = db.get_matakuliah_by_id(id)
    kelas_id = mk['kelas_id'] if mk else None
    if db.hapus_matakuliah(id):
        flash('XÃ³a mÃ´n há»c thÃ nh cÃ´ng.', 'success')
    else:
        flash('KhÃ´ng thá»ƒ xÃ³a mÃ´n há»c.', 'error')
    return redirect(url_for('kelas_detail', id=kelas_id) if kelas_id
                    else url_for('kelas_index'))


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# MANAJEMEN JADWAL
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@app.route('/jadwal')
@login_required
def jadwal_index():
    """Daftar semua jadwal."""
    return render_template('jadwal/index.html', active_page='jadwal',
                           daftar_jadwal=db.get_semua_jadwal())


@app.route('/jadwal/tambah', methods=['GET', 'POST'])
@login_required
def jadwal_tambah():
    """Tambah jadwal baru."""
    error = None
    if request.method == 'POST':
        mk_id       = request.form.get('matakuliah_id', type=int)
        hari         = request.form.get('hari', '').strip()
        jam_mulai    = request.form.get('jam_mulai', '').strip()
        jam_selesai  = request.form.get('jam_selesai', '').strip()
        if not mk_id or not hari or not jam_mulai or not jam_selesai:
            error = 'Vui lòng nhập đầy đủ thông tin.'
        elif jam_mulai >= jam_selesai:
            error = 'Giờ kết thúc phải sau giờ bắt đầu.'
        else:
            # batas_terlambat dihitung otomatis di database.py
            hasil = db.tambah_jadwal(mk_id, hari, jam_mulai, jam_selesai)
            if hasil:
                flash('Thêm lịch học thành công!', 'success')
                return redirect(url_for('jadwal_index'))
            error = 'Không thể thêm lịch học.'
    return render_template('jadwal/form.html', active_page='jadwal',
                           daftar_mk=db.get_semua_matakuliah(),
                           daftar_kelas=db.get_semua_kelas(),
                           error=error)


@app.route('/jadwal/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def jadwal_edit(id):
    """Sửa lịch học."""
    jadwal = db.get_jadwal_by_id(id)
    if not jadwal:
        flash('Lịch học không tồn tại.', 'error')
        return redirect(url_for('jadwal_index'))

    error = None
    if request.method == 'POST':
        mk_id       = request.form.get('matakuliah_id', type=int)
        hari         = request.form.get('hari', '').strip()
        jam_mulai    = request.form.get('jam_mulai', '').strip()
        jam_selesai  = request.form.get('jam_selesai', '').strip()
        if not mk_id or not hari or not jam_mulai or not jam_selesai:
            error = 'Vui lòng nhập đầy đủ thông tin.'
        elif jam_mulai >= jam_selesai:
            error = 'Giờ kết thúc phải sau giờ bắt đầu.'
        else:
            hasil = db.update_jadwal(id, mk_id, hari, jam_mulai, jam_selesai)
            if hasil:
                flash('Cập nhật lịch học thành công!', 'success')
                return redirect(url_for('jadwal_index'))
            error = 'Không thể cập nhật lịch học.'
    return render_template('jadwal/form.html', active_page='jadwal',
                           jadwal=jadwal,
                           is_edit=True,
                           daftar_mk=db.get_semua_matakuliah(),
                           daftar_kelas=db.get_semua_kelas(),
                           error=error)


@app.route('/jadwal/hapus/<int:id>', methods=['POST'])
@login_required
def jadwal_hapus(id):
    """Hapus jadwal."""
    if db.hapus_jadwal(id):
        flash('Xóa lịch học thành công.', 'success')
    else:
        flash('Không thể xóa lịch học.', 'error')
    return redirect(url_for('jadwal_index'))



# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# MANAJEMEN MAHASISWA
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _gallery_user_ids():
    """Read only the trainer diagnostics needed to label student readiness."""
    try:
        with open(FACE_GALLERY_META_PATH, 'r', encoding='utf-8') as gallery_file:
            metadata = json.load(gallery_file)
        user_ids = metadata.get('gallery_users')
        if user_ids is None:
            # Earlier gallery diagnostics exposed the same information as a
            # mapping rather than a list.
            user_ids = list((metadata.get('users') or {}).keys())
        if not isinstance(user_ids, list):
            return set()
        return {int(user_id) for user_id in user_ids}
    except (OSError, TypeError, ValueError, json.JSONDecodeError):
        return set()


def _biometric_state(student, gallery_user_ids):
    """Derive a truthful enrollment state without writing student records."""
    user_id = int(student['id'])
    folder = os.path.join(DATASET_PATH, str(user_id))
    photo_count = (
        len([name for name in os.listdir(folder) if name.lower().endswith('.jpg')])
        if os.path.isdir(folder) else 0
    )
    if user_id in gallery_user_ids:
        return photo_count, 'ready'
    if not photo_count:
        return photo_count, 'not_enrolled'
    manifest_path = os.path.join(folder, 'enrollment_manifest.json')
    try:
        with open(manifest_path, 'r', encoding='utf-8') as manifest_file:
            samples = json.load(manifest_file).get('samples')
        from face.enrollment import manifest_is_complete
        return photo_count, 'pending_gallery' if manifest_is_complete(samples) else 'incomplete'
    except (OSError, ValueError, json.JSONDecodeError):
        return photo_count, 'incomplete'


@app.route('/mahasiswa')
@login_required
def mahasiswa_index():
    """Daftar semua mahasiswa."""
    filter_kelas = request.args.get('kelas_id', type=int)
    if filter_kelas:
        daftar = db.get_users_by_kelas(filter_kelas)
    else:
        daftar = db.get_semua_user()
    gallery_user_ids = _gallery_user_ids()
    # Status chá»‰ sáºµn sÃ ng sau khi gallery thá»±c sá»± chá»©a template cá»§a sinh viÃªn.
    for m in daftar:
        m['foto_count'], m['biometric_state'] = _biometric_state(m, gallery_user_ids)
    all_students = db.get_semua_user()
    total_mhs = len(all_students)
    bio_aktif = sum(int(m['id']) in gallery_user_ids for m in all_students)
    return render_template('mahasiswa/index.html', active_page='mahasiswa',
                           daftar_mahasiswa=daftar,
                           daftar_kelas=db.get_semua_kelas(),
                           filter_kelas=filter_kelas,
                           total_mhs=total_mhs, bio_aktif=bio_aktif)


@app.route('/mahasiswa/register', methods=['GET', 'POST'])
@login_required
def mahasiswa_register():
    """Form registrasi mahasiswa baru (data + foto kamera).

    POST dari form biasa (tanpa foto) â†’ simpan data ke DB, redirect ke daftar.
    POST dari AJAX (api_foto_upload) â†’ sudah ditangani route terpisah.
    """
    error = None
    if request.method == 'POST':
        nama     = request.form.get('nama', '').strip()
        nim      = request.form.get('nim', '').strip()
        kelas_id = request.form.get('kelas_id', type=int)

        if not nama or not nim or not kelas_id:
            error = 'Vui lÃ²ng nháº­p Ä‘áº§y Ä‘á»§ há» tÃªn, mÃ£ sinh viÃªn vÃ  lá»›p.'
        elif db.nim_sudah_ada(nim):
            error = f'MÃ£ sinh viÃªn {nim} Ä‘Ã£ Ä‘Æ°á»£c Ä‘Äƒng kÃ½.'
        else:
            user_id = db.tambah_user(nama, nim, kelas_id)
            if user_id:
                flash(f'ÄÄƒng kÃ½ sinh viÃªn {nama} thÃ nh cÃ´ng! HÃ£y tiáº¿p tá»¥c chá»¥p áº£nh sinh tráº¯c há»c.', 'success')
                return redirect(url_for('mahasiswa_index'))
            else:
                error = 'KhÃ´ng thá»ƒ lÆ°u dá»¯ liá»‡u. MÃ£ sinh viÃªn cÃ³ thá»ƒ Ä‘Ã£ Ä‘Æ°á»£c sá»­ dá»¥ng.'

    from face.enrollment import ENROLLMENT_TOTAL, stages_for_total
    enrollment_total = min(FOTO_PER_USER, ENROLLMENT_TOTAL)
    return render_template(
        'mahasiswa/register.html', active_page='mahasiswa',
        daftar_kelas=db.get_semua_kelas(), error=error,
        enrollment_stages=stages_for_total(enrollment_total),
        enrollment_total=enrollment_total,
    )


@app.route('/mahasiswa/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def mahasiswa_edit(id):
    """Edit data mahasiswa."""
    mhs = db.get_user_by_id(id)
    if not mhs:
        flash('KhÃ´ng tÃ¬m tháº¥y sinh viÃªn.', 'error')
        return redirect(url_for('mahasiswa_index'))
    error = None
    if request.method == 'POST':
        nama = request.form.get('nama', '').strip()
        nim  = request.form.get('nim', '').strip()
        kid  = request.form.get('kelas_id', type=int)
        if not nama or not nim or not kid:
            error = 'Vui lÃ²ng nháº­p Ä‘áº§y Ä‘á»§ thÃ´ng tin.'
        elif db.update_user(id, nama, nim, kid):
            flash('Cáº­p nháº­t thÃ´ng tin sinh viÃªn thÃ nh cÃ´ng!', 'success')
            return redirect(url_for('mahasiswa_index'))
        else:
            error = 'KhÃ´ng thá»ƒ cáº­p nháº­t. MÃ£ sinh viÃªn cÃ³ thá»ƒ Ä‘Ã£ Ä‘Æ°á»£c sá»­ dá»¥ng.'
    return render_template('mahasiswa/edit.html', active_page='mahasiswa',
                           mhs=mhs, daftar_kelas=db.get_semua_kelas(), error=error)


@app.route('/mahasiswa/hapus/<int:id>', methods=['POST'])
@login_required
def mahasiswa_hapus(id):
    """Hapus mahasiswa + folder dataset foto."""
    mhs = db.get_user_by_id(id)
    if db.hapus_user(id):
        # Hapus folder foto dataset jika ada
        if mhs:
            folder = os.path.join(DATASET_PATH, str(id))
            if os.path.isdir(folder):
                import shutil
                shutil.rmtree(folder, ignore_errors=True)
        _start_gallery_rebuild_background()
        flash('XÃ³a sinh viÃªn thÃ nh cÃ´ng.', 'success')
    else:
        flash('KhÃ´ng thá»ƒ xÃ³a sinh viÃªn.', 'error')
    return redirect(url_for('mahasiswa_index'))


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# TAHAP 8: REKAP ABSENSI + EXPORT
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@app.route('/absensi/rekap')
@login_required
def absensi_rekap():
    """Rekap absensi dengan filter kelas, MK, dan rentang tanggal."""
    # Ambil parameter filter dari query string
    kelas_id     = request.args.get('kelas_id', type=int)
    matakuliah_id = request.args.get('matakuliah_id', type=int)
    filter_dari  = request.args.get('dari') or None
    filter_sampai = request.args.get('sampai') or None

    # Data untuk dropdown filter
    daftar_kelas = db.get_semua_kelas()
    daftar_mk    = db.get_semua_matakuliah()

    # Ambil rekap + ringkasan berdasarkan filter
    rekap    = db.get_rekap_absensi(kelas_id, filter_dari, filter_sampai, matakuliah_id)
    ringkasan = db.get_ringkasan_rekap(kelas_id, filter_dari, filter_sampai, matakuliah_id)

    # Bangun query string untuk link export (teruskan filter yang sama)
    _params = []
    if kelas_id:      _params.append(f'kelas_id={kelas_id}')
    if matakuliah_id: _params.append(f'matakuliah_id={matakuliah_id}')
    if filter_dari:   _params.append(f'dari={filter_dari}')
    if filter_sampai: _params.append(f'sampai={filter_sampai}')
    filter_query = ('&' + '&'.join(_params)) if _params else ''

    return render_template('absensi/rekap.html',
                           active_page='rekap',
                           rekap=rekap,
                           ringkasan=ringkasan,
                           daftar_kelas=daftar_kelas,
                           daftar_mk=daftar_mk,
                           filter_kelas=kelas_id,
                           filter_mk=matakuliah_id,
                           filter_dari=filter_dari,
                           filter_sampai=filter_sampai,
                           filter_query=filter_query)


@app.route('/absensi/export')
@login_required
def absensi_export():
    """Export rekap absensi ke CSV atau Excel (.xlsx).

    Query params: format=csv|xlsx, kelas_id, matakuliah_id, dari, sampai
    """
    import io
    fmt           = request.args.get('format', 'csv').lower()
    kelas_id      = request.args.get('kelas_id', type=int)
    matakuliah_id = request.args.get('matakuliah_id', type=int)
    filter_dari   = request.args.get('dari') or None
    filter_sampai = request.args.get('sampai') or None

    rekap = db.get_rekap_absensi(kelas_id, filter_dari, filter_sampai, matakuliah_id)

    # â”€â”€ Header kolom â”€â”€
    headers = ['Há» vÃ  tÃªn', 'MÃ£ sinh viÃªn', 'Lá»›p', 'MÃ´n há»c', 'MÃ£ mÃ´n há»c',
               'Thá»©', 'NgÃ y', 'Thá»i gian Ä‘iá»ƒm danh', 'Tráº¡ng thÃ¡i', 'Ghi chÃº']

    def _row(a):
        return [
            a.get('nama', ''), a.get('nim', ''), a.get('nama_kelas', ''),
            a.get('nama_mk', ''), a.get('kode_mk', ''), a.get('hari', ''),
            str(a.get('tanggal', '')), str(a.get('waktu_absen', '') or '-'),
            a.get('status', ''), a.get('alasan', '') or '-'
        ]

    timestamp = now_wib().strftime('%Y%m%d_%H%M%S')

    if fmt == 'xlsx':
        # â”€â”€ Export Excel â”€â”€
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from flask import send_file

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Tá»•ng há»£p Ä‘iá»ƒm danh'

            # Header baris 1: judul
            ws.merge_cells('A1:J1')
            ws['A1'] = 'Tá»”NG Há»¢P ÄIá»‚M DANH â€” Há»† THá»NG ÄIá»‚M DANH'
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')

            # Header baris 2: kolom
            header_fill = PatternFill('solid', fgColor='1B2024')
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=h)
                cell.font = Font(bold=True, color='8ED5FF')
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            # Data
            STATUS_COLOR = {
                'hadir': 'D1FAE5', 'terlambat': 'FEF3C7',
                'izin': 'DBEAFE', 'sakit': 'FEF3C7', 'alpha': 'FEE2E2'
            }
            for row_idx, a in enumerate(rekap, 3):
                row_data = _row(a)
                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    status = a.get('status', '')
                    # Warnai status (kolom ke-9)
                    if col_idx == 9 and status in STATUS_COLOR:
                        cell.fill = PatternFill('solid', fgColor=STATUS_COLOR[status])

            # Lebar kolom otomatis
            col_widths = [25, 15, 12, 25, 12, 10, 12, 14, 12, 25]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return send_file(buf, as_attachment=True,
                             download_name=f'rekap_absensi_{timestamp}.xlsx',
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        except ImportError:
            flash('ChÆ°a cÃ i Ä‘áº·t openpyxl. Vui lÃ²ng xuáº¥t dá»¯ liá»‡u dÆ°á»›i dáº¡ng CSV.', 'error')
            return redirect(url_for('absensi_rekap'))

    else:
        # â”€â”€ Export CSV â”€â”€
        import csv
        from flask import Response

        def generate_csv():
            buf = io.StringIO()
            writer = csv.writer(buf)
            writer.writerow(headers)
            for a in rekap:
                writer.writerow(_row(a))
            return buf.getvalue()

        return Response(
            generate_csv(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=rekap_absensi_{timestamp}.csv'}
        )


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# TAHAP 9: LAPORAN KEHADIRAN
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@app.route('/laporan')
@login_required
def laporan_index():
    """Laporan kehadiran: persentase, donut chart, ranking kelas."""
    from datetime import timedelta

    periode = request.args.get('periode', 'bulan')

    # Tentukan rentang tanggal berdasarkan periode
    today = now_wib().date()
    if periode == 'bulan':
        tanggal_dari  = today.replace(day=1).isoformat()
        tanggal_sampai = today.isoformat()
    elif periode == 'semester':
        # Semester berjalan: Januariâ€“Juni atau Juliâ€“Desember
        bulan = today.month
        if bulan <= 6:
            tanggal_dari = today.replace(month=1, day=1).isoformat()
        else:
            tanggal_dari = today.replace(month=7, day=1).isoformat()
        tanggal_sampai = today.isoformat()
    else:  # tahun
        tanggal_dari  = today.replace(month=1, day=1).isoformat()
        tanggal_sampai = today.isoformat()

    # Data laporan
    persen        = db.get_persentase_kehadiran(tanggal_dari=tanggal_dari, tanggal_sampai=tanggal_sampai)
    ranking_kelas = db.get_ranking_kelas(tanggal_dari, tanggal_sampai)
    top_mahasiswa = db.get_top_mahasiswa(tanggal_dari, tanggal_sampai)
    statistik     = db.get_statistik_dashboard(now_wib().date())

    return render_template('laporan/index.html',
                           active_page='laporan',
                           periode=periode,
                           persen=persen,
                           ranking_kelas=ranking_kelas,
                           top_mahasiswa=top_mahasiswa,
                           total_kelas=statistik.get('total_kelas', 0),
                           total_mahasiswa=statistik.get('total_mahasiswa', 0))


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# TAHAP 10: ABSENSI MANUAL ADMIN
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@app.route('/absensi/manual', methods=['GET', 'POST'])
@login_required
def absensi_manual():
    """Input / override absensi oleh admin."""
    if request.method == 'POST':
        absensi_id = request.form.get('absensi_id', type=int)
        status_baru = request.form.get('status')

        if not absensi_id or status_baru not in ('hadir', 'terlambat', 'izin', 'sakit', 'alpha'):
            flash('Dá»¯ liá»‡u khÃ´ng há»£p lá»‡.', 'error')
            return redirect(url_for('absensi_manual'))

        if db.update_status_absensi(absensi_id, status_baru):
            status_label = {
                'hadir': 'CÃ³ máº·t', 'terlambat': 'Äi muá»™n', 'izin': 'Váº¯ng cÃ³ phÃ©p',
                'sakit': 'Nghá»‰ á»‘m', 'alpha': 'Váº¯ng khÃ´ng phÃ©p'
            }.get(status_baru, status_baru)
            flash(f'ÄÃ£ Ä‘á»•i tráº¡ng thÃ¡i Ä‘iá»ƒm danh thÃ nh "{status_label}".', 'success')
        else:
            flash('KhÃ´ng thá»ƒ thay Ä‘á»•i tráº¡ng thÃ¡i Ä‘iá»ƒm danh.', 'error')
        return redirect(url_for('absensi_manual'))

    # GET â€” tampilkan daftar absensi hari ini untuk dioverride
    rekap = db.get_absensi_hari_ini(now_wib().date())
    return render_template('absensi/manual.html',
                           active_page='rekap',
                           rekap=rekap)


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# API ENDPOINTS
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@app.route('/api/absensi/hari-ini')
@login_required
def api_absensi_hari_ini():
    """Data absensi hari ini dalam format JSON."""
    try:
        data = db.get_absensi_hari_ini(now_wib().date())
        # Konversi timedelta/time ke string agar JSON-serializable
        for row in data:
            for key, val in row.items():
                if isinstance(val, timedelta):
                    total_seconds = int(val.total_seconds())
                    hours, remainder = divmod(total_seconds, 3600)
                    minutes, seconds = divmod(remainder, 60)
                    row[key] = f'{hours:02d}:{minutes:02d}:{seconds:02d}'
                elif isinstance(val, (date,)):
                    row[key] = val.isoformat()
        return jsonify({'status': 'ok', 'data': data, 'pesan': None})
    except Exception as e:
        print(f'[API] Failed to load today attendance: {e}')
        return jsonify({
            'status': 'error', 'data': [],
            'pesan': 'Không thể tải dữ liệu điểm danh.'
        }), 500


@app.route('/api/absensi/hapus/<int:absensi_id>', methods=['POST', 'DELETE'])
@login_required
def api_absensi_hapus(absensi_id):
    """Xóa lượt điểm danh hôm nay để sinh viên có thể điểm danh lại."""
    conn = None
    cursor = None
    try:
        conn = db.get_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT a.id, a.user_id, u.nama, a.status
            FROM absensi a
            JOIN users u ON a.user_id = u.id
            WHERE a.id = %s
        """, (absensi_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({
                'status': 'error',
                'pesan': 'Không tìm thấy dữ liệu điểm danh cần xóa.'
            }), 404

        cursor.execute("DELETE FROM absensi WHERE id = %s", (absensi_id,))
        conn.commit()

        return jsonify({
            'status': 'ok',
            'pesan': f"Đã xóa dữ liệu điểm danh của {row['nama']}. Sinh viên có thể điểm danh lại qua Camera.",
            'data': {'id': absensi_id, 'user_id': row['user_id'], 'nama': row['nama'], 'status': row['status']}
        })
    except Exception as e:
        if conn and conn.is_connected():
            conn.rollback()
        print(f'[API] Lỗi khi xóa điểm danh {absensi_id}: {e}')
        return jsonify({'status': 'error', 'pesan': f'Lỗi máy chủ khi xóa điểm danh: {str(e)}'}), 500
    finally:
        if cursor:
            cursor.close()
        if conn and conn.is_connected():
            conn.close()


@app.route('/api/absensi/manual', methods=['POST'])
@login_required
def api_absensi_manual():
    """Catat absensi manual oleh admin (izin, sakit, hadir, dll)."""
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        return jsonify({'status': 'error', 'pesan': 'Dá»¯ liá»‡u khÃ´ng há»£p lá»‡.'}), 400

    user_id = data.get('user_id')
    jadwal_id = data.get('jadwal_id')
    status = data.get('status', '')
    alasan = data.get('alasan', '')

    if (isinstance(user_id, bool) or isinstance(jadwal_id, bool) or
            not isinstance(user_id, int) or not isinstance(jadwal_id, int) or
            user_id <= 0 or jadwal_id <= 0 or not isinstance(status, str) or
            not isinstance(alasan, str)):
        return jsonify({'status': 'error', 'pesan': 'Dá»¯ liá»‡u khÃ´ng há»£p lá»‡.'}), 400

    status = status.strip()
    alasan = alasan.strip() or None

    if not user_id or not jadwal_id or not status:
        return jsonify({'status': 'error', 'pesan': 'Vui lÃ²ng chá»n sinh viÃªn, lá»‹ch há»c vÃ  tráº¡ng thÃ¡i.'}), 400

    if status not in ('hadir', 'terlambat', 'izin', 'sakit', 'alpha'):
        return jsonify({'status': 'error', 'pesan': 'Tráº¡ng thÃ¡i khÃ´ng há»£p lá»‡.'}), 400

    # Validasi user dan jadwal ada
    user = db.get_user_by_id(user_id)
    if not user:
        return jsonify({'status': 'error', 'pesan': 'KhÃ´ng tÃ¬m tháº¥y sinh viÃªn.'}), 404

    jadwal = db.get_jadwal_by_id(jadwal_id)
    if not jadwal:
        return jsonify({'status': 'error', 'pesan': 'KhÃ´ng tÃ¬m tháº¥y lá»‹ch há»c.'}), 404
    if int(jadwal['kelas_id']) != int(user['kelas_id']):
        return jsonify({
            'status': 'error',
            'pesan': 'Sinh viÃªn khÃ´ng thuá»™c lá»›p cá»§a lá»‹ch há»c Ä‘Ã£ chá»n.'
        }), 400
    if jadwal['hari'] != _get_nama_hari():
        return jsonify({'status': 'error', 'pesan': 'Lá»‹ch há»c Ä‘Ã£ chá»n khÃ´ng diá»…n ra hÃ´m nay.'}), 400

    tanggal = now_wib().date()
    hasil = db.catat_absensi_manual(user_id, jadwal_id, tanggal, status, alasan)

    if hasil:
        aksi = 'diperbarui' if hasil['aksi'] == 'update' else 'dicatat'
        status_label = {
            'hadir': 'CÃ³ máº·t', 'terlambat': 'Äi muá»™n', 'izin': 'Váº¯ng cÃ³ phÃ©p',
            'sakit': 'Nghá»‰ á»‘m', 'alpha': 'Váº¯ng khÃ´ng phÃ©p'
        }.get(status, status)
        aksi_label = 'cáº­p nháº­t' if aksi == 'diperbarui' else 'ghi nháº­n'
        pesan = f"ÄÃ£ {aksi_label} Ä‘iá»ƒm danh cá»§a {user['nama']} thÃ nh '{status_label}'."
        if hasil.get('status_lama'):
            status_lama_label = {
                'hadir': 'CÃ³ máº·t', 'terlambat': 'Äi muá»™n', 'izin': 'Váº¯ng cÃ³ phÃ©p',
                'sakit': 'Nghá»‰ á»‘m', 'alpha': 'Váº¯ng khÃ´ng phÃ©p'
            }.get(hasil['status_lama'], hasil['status_lama'])
            pesan += f" (trÆ°á»›c Ä‘Ã³: {status_lama_label})"
        print(f"[MANUAL] {pesan}")
        return jsonify({'status': 'ok', 'pesan': pesan, 'data': hasil})
    else:
        return jsonify({'status': 'error', 'pesan': 'KhÃ´ng thá»ƒ ghi nháº­n Ä‘iá»ƒm danh.'}), 500


@app.route('/api/mahasiswa/list')
@login_required
def api_mahasiswa_list():
    """Daftar semua mahasiswa untuk dropdown."""
    try:
        users = db.get_semua_user()
        data = [{'id': u['id'], 'nama': u['nama'], 'nim': u['nim'],
                 'kelas_id': u['kelas_id'], 'nama_kelas': u['nama_kelas']} for u in users]
        return jsonify({'status': 'ok', 'data': data, 'pesan': None})
    except Exception as e:
        print(f'[API] Failed to load student list: {e}')
        return jsonify({
            'status': 'error', 'data': [],
            'pesan': 'KhÃ´ng thá»ƒ táº£i danh sÃ¡ch sinh viÃªn.'
        }), 500


@app.route('/api/jadwal/hari-ini')
@login_required
def api_jadwal_hari_ini():
    """Daftar jadwal hari ini (semua, bukan hanya yang aktif)."""
    hari = _get_nama_hari()
    try:
        hasil = db.get_jadwal_hari(hari)
        # Konversi timedelta ke string
        for row in hasil:
            for key, val in row.items():
                if isinstance(val, timedelta):
                    total = int(val.total_seconds())
                    h, r = divmod(total, 3600)
                    m, s = divmod(r, 60)
                    row[key] = f'{h:02d}:{m:02d}'
        return jsonify({'status': 'ok', 'data': hasil})
    except Exception as e:
        print(f'[API] Failed to load today schedule: {e}')
        return jsonify({
            'status': 'error', 'data': [],
            'pesan': 'KhÃ´ng thá»ƒ táº£i lá»‹ch há»c.'
        }), 500

def _normalize_student_name(value):
    """Normalize harmless whitespace/case differences for identity checks."""
    return ' '.join(value.split()).casefold()


def _quality_enrollment_upload(nama, nim, kelas_id, frame):
    """Serialize one student's manifest transaction across tabs/requests."""
    state_key = f"{session.get('admin_id', 'anonymous')}:{nim}"
    with _enrollment_lock:
        lock_entry = _enrollment_upload_locks.setdefault(
            state_key, {'lock': threading.Lock(), 'users': 0}
        )
        lock_entry['users'] += 1
    try:
        with lock_entry['lock']:
            return _quality_enrollment_upload_locked(nama, nim, kelas_id, frame, state_key)
    except Exception as exc:
        # Camera capture is a long-lived interaction: an unavailable model or
        # transient OpenCV/filesystem failure must never turn into an HTML 500
        # that makes the browser tear down its active stream.
        from face.yolo_arcface import FaceEngineError
        if isinstance(exc, FaceEngineError):
            reason = 'face_engine_unavailable'
            message = 'Bộ nhận diện đang khởi tạo lại. Giữ nguyên camera, hệ thống sẽ tự thử lại.'
        else:
            reason = 'enrollment_processing_failed'
            message = 'Chưa thể kiểm tra ảnh này. Giữ nguyên camera, hệ thống sẽ tự thử lại.'
        print(f'[API] Enrollment upload retry ({reason}): {exc}')
        return jsonify({'status': 'retry', 'reason': reason, 'pesan': message}), 503
    finally:
        with _enrollment_lock:
            lock_entry['users'] -= 1
            if lock_entry['users'] == 0 and _enrollment_upload_locks.get(state_key) is lock_entry:
                _enrollment_upload_locks.pop(state_key, None)


def _quality_enrollment_upload_locked(nama, nim, kelas_id, frame, state_key):
    """Store one frame only after the server verifies its quality and stage."""
    from config import FOTO_PER_USER
    from face.enrollment import (
        ENROLLMENT_TOTAL, crop_detected_face, stage_for_count,
        validate_enrollment_frame,
    )
    from face.recognition import _load_engine

    user = db.get_user_by_nim(nim)
    folder = os.path.join(DATASET_PATH, str(user['id'])) if user else None
    manifest = {'schema_version': 1, 'samples': []}
    manifest_path = os.path.join(folder, 'enrollment_manifest.json') if folder else None
    if manifest_path and os.path.isfile(manifest_path):
        try:
            with open(manifest_path, 'r', encoding='utf-8') as manifest_file:
                manifest = json.load(manifest_file)
            if not isinstance(manifest.get('samples'), list):
                raise ValueError('invalid samples')
        except (OSError, ValueError, json.JSONDecodeError):
            return jsonify({
                'status': 'error', 'reason': 'enrollment_manifest_invalid',
                'pesan': 'Dá»¯ liá»‡u Ä‘Äƒng kÃ½ khuÃ´n máº·t khÃ´ng há»£p lá»‡.',
            }), 409
    accepted_count = len(manifest['samples']) if user else 0
    reset_legacy = (
        request.form.get('reset_legacy') == 'true'
        or request.args.get('reset_legacy') == 'true'
        or (request.is_json and isinstance(request.json, dict) and request.json.get('reset_legacy') is True)
    )
    if user and not os.path.isfile(manifest_path):
        legacy_count = (
            len([name for name in os.listdir(folder) if name.lower().endswith('.jpg')])
            if os.path.isdir(folder) else 0
        )
        if legacy_count:
            if reset_legacy:
                import shutil
                shutil.rmtree(folder, ignore_errors=True)
                os.makedirs(folder, exist_ok=True)
                manifest = {'schema_version': 1, 'samples': []}
                accepted_count = 0
            else:
                return jsonify({
                    'status': 'error',
                    'pesan': 'Dá»¯ liá»‡u cÅ© chÆ°a cÃ³ manifest. Báº¡n cÃ³ muá»‘n lÃ m má»›i dá»¯ liá»‡u Ä‘á»ƒ Ä‘Äƒng kÃ½ láº¡i khÃ´ng?',
                    'can_reset': True,
                }), 409
    total = min(FOTO_PER_USER, ENROLLMENT_TOTAL)
    if accepted_count >= total:
        return jsonify({'status': 'selesai', 'pesan': 'ÄÃ£ Ä‘á»§ áº£nh Ä‘Äƒng kÃ½ Ä‘áº¡t chuáº©n.'})
    stage, stage_count = stage_for_count(accepted_count)
    detector, _, _ = _load_engine()
    check = validate_enrollment_frame(frame, detector, stage['id'])
    state_now = time.monotonic()
    with _enrollment_lock:
        stale_keys = [
            key for key, value in _enrollment_states.items()
            if state_now - value.get('updated_at', state_now) > ENROLLMENT_STATE_TTL_SECONDS
        ]
        for key in stale_keys:
            _enrollment_states.pop(key, None)
        state = _enrollment_states.get(
            state_key,
            {'stage': stage['id'], 'stable': 0, 'updated_at': state_now},
        )
        if state['stage'] != stage['id']:
            state = {'stage': stage['id'], 'stable': 0, 'updated_at': state_now}
        if check.accepted:
            # The frame already passes face count, quality, pose and distance.
            # Counting valid frames is robust to harmless webcam exposure and
            # detector-box jitter; using a raw pixel fingerprint here caused
            # stable enrollment to restart even while the user held position.
            state['stable'] += 1
        else:
            state['stable'] = 0
        state['updated_at'] = state_now
        _enrollment_states[state_key] = state
        stable_count = state['stable']
    response_data = {
        'accepted': accepted_count, 'total': total, 'stage': stage['id'],
        'stage_label': stage['label'], 'stage_accepted': stage_count,
        'stage_total': stage['target'], 'stable': stable_count,
        'stable_required': ENROLLMENT_STABLE_FRAMES, 'metrics': check.metrics,
    }
    if not check.accepted:
        return jsonify({'status': 'retry', 'pesan': check.message, 'reason': check.reason, 'data': response_data})
    if stable_count < ENROLLMENT_STABLE_FRAMES:
        return jsonify({'status': 'retry', 'pesan': f'Giá»¯ yÃªn tÆ° tháº¿ ({stable_count}/{ENROLLMENT_STABLE_FRAMES})...', 'reason': 'stabilizing', 'data': response_data})

    created_user_id = None
    if not user:
        user_id = db.tambah_user(nama, nim, kelas_id)
        if not user_id:
            return jsonify({'status': 'error', 'pesan': 'KhÃ´ng thá»ƒ táº¡o sinh viÃªn.'}), 400
        created_user_id = user_id
    else:
        stored_name = _normalize_student_name(user['nama'])
        submitted_name = _normalize_student_name(nama)
        if stored_name != submitted_name or int(user['kelas_id']) != kelas_id:
            return jsonify({
                'status': 'error', 'reason': 'nim_conflict',
                'pesan': 'MÃ£ sinh viÃªn Ä‘Ã£ thuá»™c vá» sinh viÃªn khÃ¡c.',
            }), 409
        user_id = user['id']
    image_path = None
    temporary_path = None
    try:
        folder = os.path.join(DATASET_PATH, str(user_id))
        os.makedirs(folder, exist_ok=True)
        manifest_path = os.path.join(folder, 'enrollment_manifest.json')
        crop, crop_bbox = crop_detected_face(frame, check.detection)
        if crop is None or crop.size == 0:
            raise ValueError('empty enrollment face crop')
        filename = f'{accepted_count:02d}_{stage["id"]}.jpg'
        image_path = os.path.join(folder, filename)
        if not cv2.imwrite(image_path, crop):
            raise OSError('image write failed')
        manifest['samples'].append({
            'file': filename, 'stage': stage['id'], 'captured_at': now_wib().isoformat(),
            'metrics': check.metrics, 'crop_bbox': list(crop_bbox),
        })
        temporary_path = f'{manifest_path}.{uuid4().hex}.tmp'
        try:
            with open(temporary_path, 'w', encoding='utf-8') as manifest_file:
                json.dump(manifest, manifest_file, ensure_ascii=False, indent=2)
            os.replace(temporary_path, manifest_path)
        finally:
            if os.path.exists(temporary_path):
                os.remove(temporary_path)
    except Exception as exc:
        # Treat saving the crop and replacing its manifest as one transaction.
        # A retry must never advance the UI with an orphaned, unmanifested
        # image, and must never surface a raw filesystem/OpenCV error as 500.
        if image_path:
            try:
                if os.path.exists(image_path):
                    os.remove(image_path)
            except OSError as cleanup_exc:
                print(f'[API] Enrollment image rollback failed: {cleanup_exc}')
        if temporary_path:
            try:
                if os.path.exists(temporary_path):
                    os.remove(temporary_path)
            except OSError as cleanup_exc:
                print(f'[API] Enrollment manifest cleanup failed: {cleanup_exc}')
        if created_user_id:
            try:
                db.hapus_user(created_user_id)
            except Exception as cleanup_exc:
                print(f'[API] Enrollment user rollback failed: {cleanup_exc}')
        print(f'[API] Enrollment image save retry: {exc}')
        return jsonify({
            'status': 'retry', 'reason': 'enrollment_storage_failed',
            'pesan': 'Chưa thể lưu ảnh này. Giữ nguyên camera, hệ thống sẽ tự thử lại.',
        }), 503
    with _enrollment_lock:
        _enrollment_states.pop(state_key, None)
    next_count = accepted_count + 1
    next_stage, next_stage_count = stage_for_count(next_count)
    response_data.update({
        'user_id': user_id, 'accepted': next_count, 'stage_accepted': stage_count + 1,
        'next_stage': next_stage['id'] if next_stage else None,
        'next_stage_label': next_stage['label'] if next_stage else None,
        'next_stage_accepted': next_stage_count if next_stage else 0,
    })
    return jsonify({'status': 'ok', 'pesan': 'ÄÃ£ lÆ°u áº£nh Ä‘áº¡t chuáº©n.', 'data': response_data})


@app.route('/api/foto/upload', methods=['POST'])
@login_required
def api_foto_upload():
    """Terima foto wajah via AJAX (base64) dan simpan ke dataset/."""
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        return jsonify({'status': 'error', 'pesan': 'Dá»¯ liá»‡u khÃ´ng há»£p lá»‡.'}), 400

    nama_raw = data.get('nama', '')
    nim_raw = data.get('nim', '')
    kelas_id = data.get('kelas_id')
    foto_b64 = data.get('foto', '')
    index = data.get('index', 0)

    if not all(isinstance(value, str) for value in (nama_raw, nim_raw, foto_b64)):
        return jsonify({'status': 'error', 'pesan': 'Dá»¯ liá»‡u khÃ´ng há»£p lá»‡.'}), 400
    nama = nama_raw.strip()
    nim = nim_raw.strip()

    if not nama or not nim or not kelas_id or not foto_b64:
        return jsonify({'status': 'error', 'pesan': 'ThÃ´ng tin chÆ°a Ä‘áº§y Ä‘á»§.'}), 400

    try:
        index = int(index)
        kelas_id = int(kelas_id)
    except (TypeError, ValueError):
        return jsonify({'status': 'error', 'pesan': 'Chá»‰ sá»‘ áº£nh hoáº·c lá»›p khÃ´ng há»£p lá»‡.'}), 400

    try:
        from config import FOTO_PER_USER
        if index < 0 or index > FOTO_PER_USER:
            return jsonify({'status': 'error', 'pesan': 'Chá»‰ sá»‘ áº£nh khÃ´ng há»£p lá»‡.'}), 400

        # Validasi foto sepenuhnya sebelum membuat record mahasiswa baru.
        if ',' not in foto_b64:
            return jsonify({'status': 'error', 'pesan': 'Äá»‹nh dáº¡ng áº£nh khÃ´ng há»£p lá»‡.'}), 400
        _, encoded = foto_b64.split(',', 1)
        try:
            foto_bytes = base64.b64decode(encoded, validate=True)
        except (ValueError, binascii.Error):
            return jsonify({'status': 'error', 'pesan': 'Dá»¯ liá»‡u áº£nh base64 khÃ´ng há»£p lá»‡.'}), 400
        np_arr = np.frombuffer(foto_bytes, np.uint8)
        frame = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)
        if frame is None:
            return jsonify({'status': 'error', 'pesan': 'KhÃ´ng thá»ƒ Ä‘á»c tá»‡p áº£nh.'}), 400

        if data.get('protocol') != 'quality_v1':
            return jsonify({
                'status': 'error',
                'pesan': 'Giao thá»©c Ä‘Äƒng kÃ½ khuÃ´n máº·t khÃ´ng Ä‘Æ°á»£c há»— trá»£.',
            }), 400
        return _quality_enrollment_upload(nama, nim, kelas_id, frame)

    except Exception as e:
        print(f'[API] Photo upload retry: {e}')
        return jsonify({
            'status': 'retry', 'reason': 'upload_processing_failed',
            'pesan': 'Chưa thể xử lý ảnh. Giữ nguyên camera, hệ thống sẽ tự thử lại.'
        }), 503


def _start_gallery_rebuild_background(requested_user_id=None):
    """Build gallery in background if no build is running."""
    if not _training_lock.acquire(blocking=False):
        return None

    build_id = str(uuid4())
    _set_gallery_build_state(
        build_id=build_id,
        state='running',
        last_error=None,
        started_at=now_wib().isoformat(),
        finished_at=None,
        requested_user_id=requested_user_id,
    )

    def run_training():
        try:
            from face.trainer import train_model
            if not train_model():
                print('[ERROR] Gallery build finished without a usable gallery.')
                _set_gallery_build_state(
                    state='failed',
                    last_error='KhÃ´ng thá»ƒ táº¡o gallery tá»« dá»¯ liá»‡u Ä‘Äƒng kÃ½ hiá»‡n cÃ³.',
                    finished_at=now_wib().isoformat(),
                )
            else:
                _set_gallery_build_state(
                    state='succeeded',
                    last_error=None,
                    finished_at=now_wib().isoformat(),
                )
        except Exception as e:
            print(f'[ERROR] Gallery build failed: {e}')
            _set_gallery_build_state(
                state='failed',
                last_error='Cáº­p nháº­t gallery gáº·p lá»—i ná»™i bá»™. HÃ£y kiá»ƒm tra nháº­t kÃ½ mÃ¡y chá»§.',
                finished_at=now_wib().isoformat(),
            )
        finally:
            _training_lock.release()

    try:
        thread = threading.Thread(target=run_training, daemon=True)
        thread.start()
        return build_id
    except Exception:
        _set_gallery_build_state(
            build_id=build_id,
            state='failed',
            last_error='KhÃ´ng thá»ƒ khá»Ÿi cháº¡y cáº­p nháº­t gallery.',
            finished_at=now_wib().isoformat(),
        )
        _training_lock.release()
        raise


@app.route('/api/training/start', methods=['POST'])
@login_required
def api_training_start():
    """Build and hot-reload the bounded ArcFace gallery in a background thread."""

    data = request.get_json(silent=True)
    nim = data.get('nim', '').strip() if isinstance(data, dict) and isinstance(data.get('nim', ''), str) else ''
    if not nim:
        return jsonify({'status': 'error', 'pesan': 'Thiáº¿u mÃ£ sinh viÃªn.'}), 400
    user = db.get_user_by_nim(nim)
    if not user:
        return jsonify({'status': 'error', 'pesan': 'KhÃ´ng tÃ¬m tháº¥y sinh viÃªn.'}), 404
    manifest_path = os.path.join(DATASET_PATH, str(user['id']), 'enrollment_manifest.json')
    try:
        with open(manifest_path, 'r', encoding='utf-8') as manifest_file:
            manifest = json.load(manifest_file)
        from face.enrollment import manifest_is_complete, ENROLLMENT_TOTAL
        if not manifest_is_complete(manifest.get('samples'), target_total=min(FOTO_PER_USER, ENROLLMENT_TOTAL)):
            return jsonify({'status': 'error', 'pesan': f'ChÆ°a Ä‘á»§ {min(FOTO_PER_USER, ENROLLMENT_TOTAL)} áº£nh Ä‘Ãºng cÃ¡c bÆ°á»›c Ä‘Äƒng kÃ½.'}), 409
    except (OSError, ValueError, json.JSONDecodeError):
        return jsonify({'status': 'error', 'pesan': 'Dá»¯ liá»‡u Ä‘Äƒng kÃ½ khuÃ´n máº·t chÆ°a hoÃ n chá»‰nh.'}), 409

    build_id = _start_gallery_rebuild_background(requested_user_id=int(user['id']))
    if not build_id:
        current = _get_gallery_build_state()
        return jsonify({
            'status': 'error',
            'pesan': 'Gallery khuÃ´n máº·t Ä‘ang Ä‘Æ°á»£c cáº­p nháº­t. Vui lÃ²ng chá» hoÃ n táº¥t.',
            'data': {'build_id': current.get('build_id'), 'state': current.get('state')}
        }), 409

    return jsonify({
        'status': 'ok',
        'pesan': 'Da bat dau cap nhat gallery khuon mat trong nen.',
        'data': {'build_id': build_id, 'state': 'running'}
    })



@app.route('/api/training/status')
@login_required
def api_training_status():
    """Return the lifecycle of the gallery build started by enrollment."""
    build_id = request.args.get('build_id', '').strip()
    state = _get_gallery_build_state()
    if not build_id or build_id != state.get('build_id'):
        return jsonify({
            'status': 'error',
            'pesan': 'KhÃ´ng tÃ¬m tháº¥y láº§n cáº­p nháº­t gallery nÃ y.',
            'data': None,
        }), 404
    return jsonify({
        'status': 'ok',
        'pesan': 'ÄÃ£ láº¥y tráº¡ng thÃ¡i cáº­p nháº­t gallery.',
        'data': state,
    })


@app.route('/api/search')
@login_required
def api_search():
    """Endpoint API untuk mencari data mahasiswa dan jadwal secara realtime."""
    query = request.args.get('q', '').strip()
    if not query:
        return jsonify({
            'status': 'ok',
            'data': {'mahasiswa': [], 'jadwal': []},
            'pesan': 'Tá»« khÃ³a tÃ¬m kiáº¿m Ä‘ang Ä‘á»ƒ trá»‘ng.'
        })

    try:
        mahasiswa = db.cari_mahasiswa(query)
        jadwal = db.cari_jadwal(query)

        # Konversi tipe data non-JSON-serializable (seperti timedelta) ke string
        for row in jadwal:
            for key in ['jam_mulai', 'jam_selesai', 'batas_terlambat']:
                if key in row:
                    val = row[key]
                    if isinstance(val, timedelta):
                        total_seconds = int(val.total_seconds())
                        hours, remainder = divmod(total_seconds, 3600)
                        minutes, seconds = divmod(remainder, 60)
                        row[key] = f'{hours:02d}:{minutes:02d}'
                    elif hasattr(val, 'strftime'):
                        row[key] = val.strftime('%H:%M')
                    else:
                        row[key] = str(val)[:5]

        return jsonify({
            'status': 'ok',
            'data': {
                'mahasiswa': mahasiswa,
                'jadwal': jadwal
            },
            'pesan': 'TÃ¬m kiáº¿m thÃ nh cÃ´ng.'
        })
    except Exception as e:
        print(f'[API] Search failed: {e}')
        return jsonify({
            'status': 'error',
            'data': {'mahasiswa': [], 'jadwal': []},
            'pesan': 'KhÃ´ng thá»ƒ thá»±c hiá»‡n tÃ¬m kiáº¿m lÃºc nÃ y.'
        }), 500




# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# FACE RECOGNITION + ABSENSI OTOMATIS
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _decode_frame(frame_b64):
    """Decode base64 frame menjadi numpy array BGR OpenCV."""
    try:
        # Hapus header data URI jika ada
        if ',' in frame_b64:
            frame_b64 = frame_b64.split(',', 1)[1]
        img_bytes = base64.b64decode(frame_b64)
        np_arr = np.frombuffer(img_bytes, np.uint8)
        frame = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)
        return frame
    except Exception as e:
        print(f'[ERROR] Gagal decode frame: {e}')
        return None


def _simpan_snapshot(frame, user_id):
    """Simpan snapshot bukti absensi ke folder snapshots/."""
    try:
        os.makedirs(SNAPSHOT_PATH, exist_ok=True)
        now = now_wib()
        filename = (
            f"{user_id}_{now.strftime('%Y%m%d_%H%M%S_%f')}_"
            f"{uuid4().hex[:8]}.jpg"
        )
        filepath = os.path.join(SNAPSHOT_PATH, filename)
        cv2.imwrite(filepath, frame)
        return filepath
    except Exception as e:
        print(f'[ERROR] Gagal simpan snapshot: {e}')
        return None


def _hapus_snapshot_gagal(snapshot_path):
    """XÃ³a snapshot cá»§a insert tháº¥t báº¡i, chá»‰ khi file náº±m trong SNAPSHOT_PATH."""
    if not snapshot_path:
        return
    try:
        base_path = os.path.realpath(SNAPSHOT_PATH)
        target_path = os.path.realpath(snapshot_path)
        if os.path.commonpath([base_path, target_path]) != base_path:
            return
        if os.path.isfile(target_path):
            os.remove(target_path)
    except (OSError, ValueError) as error:
        print(f'[SNAPSHOT] Gagal membersihkan snapshot orphan: {error}')


def _kirim_ke_esp32(nama, nim, status_pesan):
    """Kirim notifikasi ke ESP32 via HTTP POST (jika diaktifkan)."""
    if not ESP32_ENABLED:
        return
    try:
        import requests
        url = f"http://{ESP32_IP}:{ESP32_PORT}/absensi"
        payload = {'nama': nama, 'nim': nim, 'status': status_pesan}
        requests.post(url, json=payload, timeout=ESP32_TIMEOUT)
        print(f'[ESP32] Notifikasi terkirim: {nama} - {status_pesan}')
    except Exception as e:
        print(f'[ESP32] Gagal kirim: {e}')


def _get_nama_hari():
    """Dapatkan nama hari ini dalam Bahasa Indonesia (timezone WIB)."""
    hari_map = {
        0: 'Senin', 1: 'Selasa', 2: 'Rabu',
        3: 'Kamis', 4: 'Jumat', 5: 'Sabtu', 6: 'Minggu'
    }
    return hari_map.get(now_wib().weekday(), '')


def _get_ten_thu_hien_thi(nama_hari):
    """Ãnh xáº¡ tÃªn ngÃ y ná»™i bá»™ sang tiáº¿ng Viá»‡t chá»‰ táº¡i táº§ng hiá»ƒn thá»‹."""
    return {
        'Senin': 'Thá»© Hai', 'Selasa': 'Thá»© Ba', 'Rabu': 'Thá»© TÆ°',
        'Kamis': 'Thá»© NÄƒm', 'Jumat': 'Thá»© SÃ¡u', 'Sabtu': 'Thá»© Báº£y',
        'Minggu': 'Chá»§ Nháº­t'
    }.get(nama_hari, nama_hari)


def _bbox_iou(first_bbox, second_bbox):
    """TÃ­nh Intersection over Union cá»§a hai bbox `(x, y, w, h)`."""
    if first_bbox is None or second_bbox is None:
        return 0.0
    ax, ay, aw, ah = (int(value) for value in first_bbox)
    bx, by, bw, bh = (int(value) for value in second_bbox)
    left, top = max(ax, bx), max(ay, by)
    right, bottom = min(ax + aw, bx + bw), min(ay + ah, by + bh)
    intersection = max(0, right - left) * max(0, bottom - top)
    union = max(0, aw * ah) + max(0, bw * bh) - intersection
    return intersection / union if union > 0 else 0.0


def _valid_camera_client_id(client_id):
    return bool(re.fullmatch(r'[A-Za-z0-9_-]{8,64}', str(client_id or '').strip()))


def _scanner_tracker_key(client_id):
    return f'browser:{session.get("admin_id")}:{str(client_id).strip()}'


def _db_call_strict(func, *args, **kwargs):
    if hasattr(func, 'mock_calls'):
        return func(*args, **kwargs)
    try:
        return func(*args, **kwargs, raise_on_error=True)
    except TypeError as exc:
        if 'raise_on_error' not in str(exc):
            raise
        return func(*args, **kwargs)


def _database_unavailable_response(spoof_result=None):
    return {
        'status': 'error',
        'tipe': 'database_unavailable',
        'pesan': 'KhÃ´ng thá»ƒ Ä‘á»c/ghi cÆ¡ sá»Ÿ dá»¯ liá»‡u lÃºc nÃ y. HÃ£y thá»­ láº¡i sau.',
        'spoofing': spoof_result or {'is_real': True, 'label': 'DEFERRED', 'score': None}
    }


def _select_active_schedule_for_user(user, jadwal_list, spoof_result):
    matching = [
        item for item in jadwal_list
        if int(item.get('kelas_id', -1)) == int(user['kelas_id'])
    ]
    identity = {'nama': user['nama'], 'nim': user['nim']}
    if not matching:
        return None, {
            'status': 'error',
            'tipe': 'no_jadwal',
            'pesan': f'Hiá»‡n khÃ´ng cÃ³ lá»‹ch há»c cho lá»›p {user.get("nama_kelas", "")}.',
            'data': identity,
            'spoofing': spoof_result
        }
    if len(matching) > 1:
        now_str = now_wib().strftime('%H:%M:%S')
        strict = [
            item for item in matching
            if item.get('jam_mulai') and item.get('jam_selesai')
            and str(item['jam_mulai']) <= now_str <= str(item['jam_selesai'])
        ]
        if len(strict) == 1:
            return strict[0], None

        return None, {
            'status': 'error',
            'tipe': 'multiple_active_schedules',
            'pesan': 'CÃ³ nhiá»u lá»‹ch há»c Ä‘ang diá»…n ra cho cÃ¹ng lá»›p. Vui lÃ²ng kiá»ƒm tra láº¡i lá»‹ch há»c.',
            'data': identity,
            'spoofing': spoof_result
        }
    return matching[0], None


def _make_terminal_duplicate_response(cached):
    data = dict(cached.get('data') or {})
    return {
        'status': 'error',
        'tipe': 'duplikat',
        'pesan': cached.get('pesan') or 'Sinh vien da diem danh trong buoi nay.',
        'data': data,
        'spoofing': {'is_real': True, 'label': 'CACHED', 'score': None},
        'cached': True
    }


def _mark_completed_track(tracker_key, track_id, user_id, jadwal_id, response):
    if track_id is None or user_id is None or jadwal_id is None:
        return
    camera_key = str(tracker_key)
    key = (camera_key, int(track_id))
    with _consecutive_lock:
        _completed_trackers[key] = {
            'user_id': int(user_id),
            'jadwal_id': int(jadwal_id),
            'bbox': response.get('bbox'),
            'data': dict(response.get('data') or {}),
            'pesan': response.get('pesan'),
            'updated_at': time.monotonic(),
        }
    try:
        from face.recognition import mark_track_completed
        data = response.get('data') or {}
        mark_track_completed(
            tracker_key, track_id, user_id,
            confidence=data.get('confidence'),
            match_score=response.get('match_score') or data.get('confidence'),
        )
    except Exception as exc:
        print(f'[ABSENSI] Gagal pin completed track: {exc}')


def _get_completed_track_response(tracker_key, track_id, user_id):
    if track_id is None or user_id is None:
        return None
    key = (str(tracker_key), int(track_id))
    with _consecutive_lock:
        cached = _completed_trackers.get(key)
        if not cached:
            return None
        if time.monotonic() - cached.get('updated_at', 0) > 300:
            _completed_trackers.pop(key, None)
            return None
        if cached.get('user_id') != int(user_id):
            _completed_trackers.pop(key, None)
            return None
        cached['updated_at'] = time.monotonic()
        return _make_terminal_duplicate_response(cached)


def _sync_face_trackers(tracker_key, detections):
    """Count consecutive evidence for the same track, identity and position."""
    normalized = {}
    for track_id, payload in detections.items():
        if track_id is None:
            continue
        if isinstance(payload, dict):
            bbox = payload.get('bbox')
            user_id = payload.get('user_id')
        else:
            # Backward compatibility for the legacy single-face helpers/tests.
            bbox = payload
            user_id = track_id
        normalized[int(track_id)] = {
            'bbox': None if bbox is None else tuple(int(value) for value in bbox),
            'user_id': None if user_id is None else int(user_id),
        }
    camera_key = str(tracker_key)

    with _consecutive_lock:
        now_monotonic = time.monotonic()
        stale_keys = [
            key for key, value in _consecutive_trackers.items()
            if now_monotonic - value.get('updated_at', now_monotonic) > 300
        ]
        for key in stale_keys:
            _consecutive_trackers.pop(key, None)
            _completed_trackers.pop(key, None)
        completed_stale_keys = [
            key for key, value in _completed_trackers.items()
            if now_monotonic - value.get('updated_at', now_monotonic) > 300
        ]
        for key in completed_stale_keys:
            _completed_trackers.pop(key, None)

        camera_entries = [
            key for key in _consecutive_trackers
            if key[0] == camera_key
        ]
        for key in camera_entries:
            track_id = key[1]
            tracker = _consecutive_trackers[key]
            if track_id in normalized:
                current = normalized[track_id]
                previous_bbox = tracker.get('bbox')
                current_bbox = current['bbox']
                same_position = (
                    previous_bbox is None and current_bbox is None
                ) or _bbox_iou(previous_bbox, current_bbox) >= 0.25
                same_identity = tracker.get('user_id') == current['user_id']
                tracker['count'] = (
                    tracker['count'] + 1 if same_position and same_identity else 1
                )
                tracker['bbox'] = current_bbox
                tracker['user_id'] = current['user_id']
                tracker['updated_at'] = now_monotonic
            else:
                # The evidence is consecutive: a missing/unknown/low-quality
                # frame cannot carry verification credit into a later frame.
                _consecutive_trackers.pop(key, None)
                _completed_trackers.pop(key, None)

        completed_entries = [
            key for key in _completed_trackers
            if key[0] == camera_key and key[1] not in normalized
        ]
        for key in completed_entries:
            _completed_trackers.pop(key, None)

        for track_id, current in normalized.items():
            key = (camera_key, track_id)
            if key not in _consecutive_trackers:
                _consecutive_trackers[key] = {
                    'count': 1,
                    'bbox': current['bbox'],
                    'user_id': current['user_id'],
                    'updated_at': now_monotonic
                }

        return {
            track_id: _consecutive_trackers[(camera_key, track_id)]['count']
            for track_id in normalized
        }


def _sync_consecutive_trackers(tracker_key, detected_uids):
    """API tÆ°Æ¡ng thÃ­ch cho caller má»™t khuÃ´n máº·t khÃ´ng cung cáº¥p bbox."""
    return _sync_face_trackers(
        tracker_key,
        {int(uid): None for uid in detected_uids if uid is not None}
    )


def _update_consecutive_tracker(tracker_key, detected_uid=None):
    """API tÆ°Æ¡ng thÃ­ch cho luá»“ng/test má»™t khuÃ´n máº·t."""
    counts = _sync_consecutive_trackers(
        tracker_key, [] if detected_uid is None else [detected_uid]
    )
    if detected_uid is None:
        return 0
    return counts[int(detected_uid)]


def _reset_consecutive_tracker(tracker_key, user_id=None, remove=False):
    """XÃ³a bá»™ Ä‘áº¿m cá»§a má»™t ngÆ°á»i hoáº·c toÃ n bá»™ camera/client."""
    camera_key = str(tracker_key)
    with _consecutive_lock:
        keys = [
            key for key in _consecutive_trackers
            if key[0] == camera_key and (user_id is None or key[1] == int(user_id))
        ]
        for key in keys:
            _consecutive_trackers.pop(key, None)
            _completed_trackers.pop(key, None)
    if remove and user_id is None:
        from face.recognition import reset_tracker
        reset_tracker(camera_key)


def _proses_recognition_single(frame, tracker_key='default'):
    """Proses satu frame: anti-spoofing â†’ recognition â†’ absensi.

    Alur lengkap sesuai context.md bagian 5.4:
    1. Cek anti-spoofing
    2. Predict wajah báº±ng gallery ArcFace
    3. Cari jadwal aktif hari ini
    4. Cek duplikasi absensi
    5. Tentukan status (hadir/terlambat)
    6. Simpan snapshot + catat absensi
    7. Kirim ke ESP32

    Returns:
        dict hasil proses untuk dikirim ke client
    """
    from face.anti_spoofing import check as spoofing_check
    from face.recognition import predict_single

    # â”€â”€ 1. Anti-spoofing â”€â”€
    spoof_result = spoofing_check(frame)

    if not spoof_result['is_real']:
        # Spoofing terdeteksi â€” simpan bukti ke spoofing_log
        snapshot = _simpan_snapshot(frame, 'spoofing')
        db.catat_spoofing(snapshot, spoof_result['score'])
        return {
            'status': 'error',
            'tipe': 'spoofing',
            'score': spoof_result['score'],
            'pesan': 'PhÃ¡t hiá»‡n hÃ nh vi giáº£ máº¡o! Vui lÃ²ng sá»­ dá»¥ng khuÃ´n máº·t tháº­t.',
            'spoofing': spoof_result
        }

    if spoof_result['label'] == 'NO_FACE':
        return {
            'status': 'skip',
            'tipe': 'no_face',
            'pesan': 'KhÃ´ng phÃ¡t hiá»‡n tháº¥y khuÃ´n máº·t.',
            'spoofing': spoof_result
        }

    # â”€â”€ 2. Predict wajah báº±ng gallery ArcFace â”€â”€
    result = predict_single(frame)

    if result is None:
        return {
            'status': 'skip',
            'tipe': 'no_face',
            'pesan': 'KhÃ´ng phÃ¡t hiá»‡n tháº¥y khuÃ´n máº·t Ä‘á»ƒ nháº­n diá»‡n.',
            'spoofing': spoof_result
        }

    # Log confidence untuk debug
    print(f'[ABSENSI] Predict: user_id={result["user_id"]}, '
          f'conf={result["confidence"]:.1f}, dikenali={result["dikenali"]}')

    if not result['dikenali']:
        # Kurangi counter saja (toleransi 1 frame buruk), tidak reset total
        _update_consecutive_tracker(tracker_key)
        return {
            'status': 'error',
            'tipe': 'unknown',
            'confidence': result['confidence'],
            'pesan': f'KhÃ´ng nháº­n diá»‡n Ä‘Æ°á»£c khuÃ´n máº·t (Ä‘á»™ tin cáº­y: {result["confidence"]}).',
            'spoofing': spoof_result
        }

    # â”€â”€ Verifikasi konsekutif: harus 3x berturut-turut user yang SAMA â”€â”€
    detected_uid = result['user_id']
    verification_count = _update_consecutive_tracker(tracker_key, detected_uid)
    required = RECOGNITION_REQUIRED_FRAMES
    if verification_count < required:
        return {
            'status': 'skip',
            'tipe': 'verifying',
            'pesan': f'Äang xÃ¡c minh khuÃ´n máº·t... ({verification_count}/{required})',
            'spoofing': spoof_result
        }

    # Reset counter setelah berhasil verifikasi
    _reset_consecutive_tracker(tracker_key)

    user_id = result['user_id']
    confidence = result['confidence']

    # â”€â”€ 3. Ambil data mahasiswa â”€â”€
    user = db.get_user_by_id(user_id)
    if not user:
        return {
            'status': 'error',
            'tipe': 'user_not_found',
            'pesan': f'KhÃ´ng tÃ¬m tháº¥y ngÆ°á»i dÃ¹ng cÃ³ ID {user_id} trong cÆ¡ sá»Ÿ dá»¯ liá»‡u.',
            'spoofing': spoof_result
        }

    # â”€â”€ 4. Cari jadwal aktif hari ini â”€â”€
    hari = _get_nama_hari()
    waktu_sekarang = now_wib().strftime('%H:%M:%S')
    jadwal_list = db.get_jadwal_aktif(hari, waktu_sekarang)

    if not jadwal_list:
        return {
            'status': 'error',
            'tipe': 'no_jadwal',
            'pesan': f'Hiá»‡n khÃ´ng cÃ³ lá»‹ch há»c nÃ o Ä‘ang diá»…n ra ({_get_ten_thu_hien_thi(hari)} {waktu_sekarang}).',
            'data': {'nama': user['nama'], 'nim': user['nim']},
            'spoofing': spoof_result
        }

    jadwal, schedule_error = _select_active_schedule_for_user(
        user, jadwal_list, spoof_result
    )
    if schedule_error:
        return schedule_error

    # â”€â”€ 5. Cek duplikasi absensi â”€â”€
    tanggal_hari_ini = now_wib().date()
    sudah = _db_call_strict(
        db.cek_sudah_absen, user_id, jadwal['id'], tanggal_hari_ini
    )

    if sudah:
        # Kirim notifikasi duplikat ke ESP32
        _kirim_ke_esp32(user['nama'], user['nim'], 'duplikat')
        return {
            'status': 'error',
            'tipe': 'duplikat',
            'pesan': f'{user["nama"]} Ä‘Ã£ Ä‘iá»ƒm danh mÃ´n {jadwal["nama_mk"]} hÃ´m nay.',
            'data': {
                'nama': user['nama'], 'nim': user['nim'],
                'status_absensi': sudah['status'],
                'jadwal_id': jadwal['id']
            },
            'spoofing': spoof_result
        }

    # â”€â”€ 6. Tentukan status: hadir atau terlambat â”€â”€
    batas_str = str(jadwal['batas_terlambat'])
    # Konversi timedelta ke string waktu jika perlu
    if isinstance(jadwal['batas_terlambat'], timedelta):
        total_sec = int(jadwal['batas_terlambat'].total_seconds())
        h, m, s = total_sec // 3600, (total_sec % 3600) // 60, total_sec % 60
        batas_str = f'{h:02d}:{m:02d}:{s:02d}'

    status_absensi = 'hadir' if waktu_sekarang <= batas_str else 'terlambat'

    # â”€â”€ 7. Simpan snapshot bukti absensi â”€â”€
    snapshot_path = _simpan_snapshot(frame, user_id)

    # â”€â”€ 8. Catat absensi ke database â”€â”€
    try:
        absensi_id = _db_call_strict(
            db.catat_absensi,
            user_id=user_id,
            jadwal_id=jadwal['id'],
            tanggal=tanggal_hari_ini,
            waktu_absen=waktu_sekarang,
            status=status_absensi,
            snapshot_path=snapshot_path,
            dibuat_manual=False
        )
    except db.DatabaseQueryError:
        _hapus_snapshot_gagal(snapshot_path)
        raise

    if not absensi_id:
        _hapus_snapshot_gagal(snapshot_path)
        sudah_setelah_insert = _db_call_strict(
            db.cek_sudah_absen, user_id, jadwal['id'], tanggal_hari_ini
        )
        if sudah_setelah_insert:
            return {
                'status': 'error',
                'tipe': 'duplikat',
                'pesan': f'{user["nama"]} Ä‘Ã£ Ä‘iá»ƒm danh mÃ´n {jadwal["nama_mk"]} hÃ´m nay.',
                'data': {
                    'nama': user['nama'], 'nim': user['nim'],
                    'status_absensi': sudah_setelah_insert['status'],
                    'jadwal_id': jadwal['id']
                },
                'spoofing': spoof_result
            }
        return {
            'status': 'error',
            'tipe': 'db_error',
            'pesan': 'KhÃ´ng thá»ƒ lÆ°u dá»¯ liá»‡u Ä‘iá»ƒm danh vÃ o cÆ¡ sá»Ÿ dá»¯ liá»‡u.',
            'spoofing': spoof_result
        }

    # â”€â”€ 9. Kirim ke ESP32 â”€â”€
    esp_status = 'berhasil'
    _kirim_ke_esp32(user['nama'], user['nim'], esp_status)

    # â”€â”€ 10. Siapkan response â”€â”€
    data_response = {
        'nama': user['nama'],
        'nim': user['nim'],
        'nama_kelas': user.get('nama_kelas', ''),
        'nama_mk': jadwal['nama_mk'],
        'jadwal_id': jadwal['id'],
        'confidence': confidence,
        'status_absensi': status_absensi,
        'waktu_absen': waktu_sekarang,
        'absensi_id': absensi_id,
        'status': status_absensi
    }

    # Ambil statistik terbaru untuk update dashboard
    stats = db.get_statistik_dashboard(tanggal_hari_ini)

    return {
        'status': 'ok',
        'pesan': f'Äiá»ƒm danh cho {user["nama"]} thÃ nh cÃ´ng.',
        'data': data_response,
        'stats': {
            'hadir': stats.get('hadir_hari_ini', 0),
            'terlambat': stats.get('terlambat_hari_ini', 0),
            'alpha': stats.get('alpha_hari_ini', 0)
        },
        'spoofing': spoof_result
    }


def _process_verified_prediction(
    frame, prediction, spoof_result, hari, waktu_sekarang,
    tanggal_hari_ini, jadwal_list
):
    """Ghi Ä‘iá»ƒm danh cho má»™t danh tÃ­nh Ä‘Ã£ Ä‘á»§ sá»‘ frame xÃ¡c nháº­n."""
    user_id = int(prediction['user_id'])
    confidence = prediction['confidence']
    user = _db_call_strict(db.get_user_by_id, user_id)
    if not user:
        return {
            'status': 'error',
            'tipe': 'user_not_found',
            'pesan': f'KhÃ´ng tÃ¬m tháº¥y ngÆ°á»i dÃ¹ng cÃ³ ID {user_id} trong cÆ¡ sá»Ÿ dá»¯ liá»‡u.',
            'spoofing': spoof_result
        }

    if not jadwal_list:
        return {
            'status': 'error',
            'tipe': 'no_jadwal',
            'pesan': f'Hiá»‡n khÃ´ng cÃ³ lá»‹ch há»c nÃ o Ä‘ang diá»…n ra ({_get_ten_thu_hien_thi(hari)} {waktu_sekarang}).',
            'data': {'nama': user['nama'], 'nim': user['nim']},
            'spoofing': spoof_result
        }

    jadwal, schedule_error = _select_active_schedule_for_user(
        user, jadwal_list, spoof_result
    )
    if schedule_error:
        return schedule_error

    sudah = _db_call_strict(
        db.cek_sudah_absen, user_id, jadwal['id'], tanggal_hari_ini
    )
    if sudah:
        _kirim_ke_esp32(user['nama'], user['nim'], 'duplikat')
        return {
            'status': 'error',
            'tipe': 'duplikat',
            'pesan': f'{user["nama"]} Ä‘Ã£ Ä‘iá»ƒm danh mÃ´n {jadwal["nama_mk"]} hÃ´m nay.',
            'data': {
                'nama': user['nama'],
                'nim': user['nim'],
                'status_absensi': sudah['status'],
                'jadwal_id': jadwal['id']
            },
            'spoofing': spoof_result
        }

    batas_str = str(jadwal['batas_terlambat'])
    if isinstance(jadwal['batas_terlambat'], timedelta):
        total_sec = int(jadwal['batas_terlambat'].total_seconds())
        h, m, s = total_sec // 3600, (total_sec % 3600) // 60, total_sec % 60
        batas_str = f'{h:02d}:{m:02d}:{s:02d}'
    status_absensi = 'hadir' if waktu_sekarang <= batas_str else 'terlambat'

    snapshot_path = _simpan_snapshot(frame, user_id)
    try:
        absensi_id = _db_call_strict(
            db.catat_absensi,
            user_id=user_id,
            jadwal_id=jadwal['id'],
            tanggal=tanggal_hari_ini,
            waktu_absen=waktu_sekarang,
            status=status_absensi,
            snapshot_path=snapshot_path,
            dibuat_manual=False
        )
    except db.DatabaseQueryError:
        _hapus_snapshot_gagal(snapshot_path)
        raise
    if not absensi_id:
        _hapus_snapshot_gagal(snapshot_path)
        sudah_setelah_insert = _db_call_strict(
            db.cek_sudah_absen,
            user_id, jadwal['id'], tanggal_hari_ini
        )
        if sudah_setelah_insert:
            return {
                'status': 'error',
                'tipe': 'duplikat',
                'pesan': f'{user["nama"]} Ä‘Ã£ Ä‘iá»ƒm danh mÃ´n {jadwal["nama_mk"]} hÃ´m nay.',
                'data': {
                    'nama': user['nama'],
                    'nim': user['nim'],
                    'status_absensi': sudah_setelah_insert['status'],
                    'jadwal_id': jadwal['id']
                },
                'spoofing': spoof_result
            }
        return {
            'status': 'error',
            'tipe': 'db_error',
            'pesan': 'KhÃ´ng thá»ƒ lÆ°u dá»¯ liá»‡u Ä‘iá»ƒm danh vÃ o cÆ¡ sá»Ÿ dá»¯ liá»‡u.',
            'spoofing': spoof_result
        }

    _kirim_ke_esp32(user['nama'], user['nim'], 'berhasil')
    return {
        'status': 'ok',
        'pesan': f'Äiá»ƒm danh cho {user["nama"]} thÃ nh cÃ´ng.',
        'data': {
            'nama': user['nama'],
            'nim': user['nim'],
            'nama_kelas': user.get('nama_kelas', ''),
            'nama_mk': jadwal['nama_mk'],
            'jadwal_id': jadwal['id'],
            'confidence': confidence,
            'status_absensi': status_absensi,
            'waktu_absen': waktu_sekarang,
            'absensi_id': absensi_id,
            'status': status_absensi
        },
        'spoofing': spoof_result
    }


def _attach_face_metadata(response, prediction, face_index):
    """Gáº¯n metadata bbox vÃ o má»™t káº¿t quáº£ nhÆ°ng khÃ´ng lÃ m thay Ä‘á»•i object gá»‘c."""
    enriched = dict(response)
    enriched['face_index'] = face_index
    enriched['bbox'] = [int(value) for value in prediction.get('bbox', (0, 0, 0, 0))]
    enriched['track_id'] = prediction.get('track_id')
    enriched['detector_score'] = prediction.get('detector_score')
    enriched['quality_reason'] = prediction.get('quality_reason')
    enriched['pipeline_latency_ms'] = prediction.get('pipeline_latency_ms')
    enriched.setdefault('confidence', prediction.get('confidence'))
    enriched.setdefault('user_id', prediction.get('user_id'))
    enriched['match_score'] = prediction.get('match_score')

    tipe = enriched.get('tipe')
    if tipe in ('spoofing', 'identity_conflict', 'unknown', 'low_quality'):
        display_status = 'error'
    elif tipe in (
        'verifying', 'no_jadwal',
        'multiple_active_schedules', 'database_unavailable',
        'needs_calibration'
    ):
        display_status = 'warning'
    else:
        display_status = 'recognized'

    labels = {
        'spoofing': 'Giả mạo',
        'identity_conflict': 'Xung đột danh tính',
        'unknown': 'Không khớp',
        'no_jadwal': 'Đã nhận diện — không có lịch học',
        'needs_calibration': 'Cần hiệu chuẩn',
        'duplikat': 'Đã điểm danh',
    }
    if tipe == 'low_quality':
        quality_labels = {
            'face_too_small': 'Lại gần máy ảnh hơn',
            'face_too_dark': 'Tăng ánh sáng khuôn mặt',
            'face_too_bright': 'Giảm ánh sáng gắt',
            'face_blurry': 'Giữ yên khuôn mặt',
            'landmarks_invalid': 'Nhìn thẳng vào máy ảnh',
            'face_out_of_box': 'Đặt trọn khuôn mặt vào khung',
            'face_out_of_frame': 'Đặt trọn khuôn mặt vào khung',
        }
        label = quality_labels.get(prediction.get('quality_reason'), 'Cần điều chỉnh khuôn mặt')
    elif tipe in ('verifying', 'verification_progress'):
        verification_count = int(enriched.get('verification_count') or 0)
        required_frames = int(enriched.get('required_frames') or RECOGNITION_REQUIRED_FRAMES)
        label = f'Đang xác minh ({verification_count}/{required_frames})'
    elif enriched.get('data') and (
        enriched.get('status') == 'ok' or tipe == 'duplikat'
    ):
        data = enriched['data']
        identity_parts = [
            str(value).strip()
            for value in (data.get('nama'), data.get('nim'))
            if value is not None and str(value).strip()
        ]
        label = ' — '.join(identity_parts) or 'Đã nhận diện'
    else:
        label = labels.get(tipe, 'Đã nhận diện')
    if tipe == 'multiple_active_schedules':
        label = 'Lịch học bị trùng'
    elif tipe == 'database_unavailable':
        label = 'CSDL chua san sang'
    enriched['display_status'] = display_status
    enriched['display_label'] = label
    return enriched


def _aggregate_face_results(face_results, tanggal_hari_ini):
    """Táº¡o response tÆ°Æ¡ng thÃ­ch API cÅ© vÃ  bá»• sung danh sÃ¡ch káº¿t quáº£ nhiá»u máº·t."""
    successes = [item for item in face_results if item.get('status') == 'ok']
    verifying = [
        item for item in face_results
        if item.get('status') == 'skip' and item.get('tipe') == 'verifying'
    ]
    priority = (
        successes or verifying or
        [item for item in face_results if item.get('tipe') == 'duplikat'] or
        [item for item in face_results if item.get('tipe') == 'spoofing'] or
        [item for item in face_results if item.get('tipe') == 'unknown'] or
        face_results
    )
    response = dict(priority[0])
    response['results'] = face_results
    response['summary'] = {
        'total_faces': len(face_results),
        'recorded': len(successes),
        'verifying': len(verifying),
        'duplicates': sum(item.get('tipe') == 'duplikat' for item in face_results),
        'unknown': sum(item.get('tipe') == 'unknown' for item in face_results),
        'low_quality': sum(item.get('tipe') == 'low_quality' for item in face_results),
        'database_unavailable': sum(
            item.get('tipe') == 'database_unavailable' for item in face_results
        ),
        'multiple_active_schedules': sum(
            item.get('tipe') == 'multiple_active_schedules' for item in face_results
        ),
        'needs_calibration': sum(
            item.get('tipe') == 'needs_calibration' for item in face_results
        ),
        'spoofing': sum(item.get('tipe') == 'spoofing' for item in face_results),
        'identity_conflicts': sum(
            item.get('tipe') == 'identity_conflict' for item in face_results
        ),
    }
    if successes:
        stats = db.get_statistik_dashboard(tanggal_hari_ini)
        response['stats'] = {
            'hadir': stats.get('hadir_hari_ini', 0),
            'terlambat': stats.get('terlambat_hari_ini', 0),
            'alpha': stats.get('alpha_hari_ini', 0)
        }
        if len(successes) > 1:
            response['pesan'] = f'ÄÃ£ ghi Ä‘iá»ƒm danh cho {len(successes)} sinh viÃªn.'
    return response


def _successful_face_results(result):
    """Láº¥y má»i káº¿t quáº£ vá»«a ghi thÃ nh cÃ´ng, ká»ƒ cáº£ response má»™t khuÃ´n máº·t cÅ©."""
    successes = [
        item for item in result.get('results', [])
        if item.get('status') == 'ok' and item.get('data')
    ]
    if not successes and result.get('status') == 'ok' and result.get('data'):
        successes = [result]
    return successes


def _broadcast_absensi_updates(result, skip_sid=None):
    """Broadcast tá»«ng lÆ°á»£t Ä‘iá»ƒm danh, dÃ¹ng chung cho WebSocket vÃ  HTTP fallback."""
    emit_options = {'skip_sid': skip_sid} if skip_sid else {}
    for item in _successful_face_results(result):
        socketio.emit('absensi_update', {
            **item['data'],
            'stats': result.get('stats')
        }, **emit_options)


def _proses_recognition(frame, tracker_key='default'):
    """Nháº­n diá»‡n vÃ  xá»­ lÃ½ Ä‘iá»ƒm danh cho táº¥t cáº£ khuÃ´n máº·t trong má»™t frame."""
    from face.anti_spoofing import check_face
    from face.recognition import predict

    predictions = list(predict(frame, tracker_key=tracker_key) or [])
    if not predictions:
        _sync_consecutive_trackers(tracker_key, [])
        return {
            'status': 'skip',
            'tipe': 'no_face',
            'pesan': 'KhÃ´ng phÃ¡t hiá»‡n tháº¥y khuÃ´n máº·t.',
            'results': [],
            'summary': {
                'total_faces': 0, 'recorded': 0, 'verifying': 0,
                'duplicates': 0, 'unknown': 0, 'needs_calibration': 0, 'spoofing': 0,
                'identity_conflicts': 0
            }
        }

    # Anti-spoofing pháº£i cháº¡y theo tá»«ng bbox, khÃ´ng dÃ¹ng káº¿t quáº£ cá»§a máº·t lá»›n nháº¥t
    # cho toÃ n bá»™ nhÃ³m ngÆ°á»i trong frame.
    # Liveness runs only when a recognized track reaches confirmation, not on
    # every detected face. This keeps detector boxes responsive.
    prepared = [(face_index, prediction, None)
                for face_index, prediction in enumerate(predictions)]

    # Náº¿u nhiá»u bbox cÃ¹ng bá»‹ gÃ¡n má»™t user_id thÃ¬ khÃ´ng bbox nÃ o Ä‘Æ°á»£c phÃ©p ghi
    # Ä‘iá»ƒm danh: Ä‘Ã¢y lÃ  xung Ä‘á»™t danh tÃ­nh, khÃ´ng pháº£i cÄƒn cá»© Ä‘á»ƒ chá»n má»™t máº·t.
    candidates_by_user = {}
    for face_index, prediction, spoof_result in prepared:
        if not prediction.get('dikenali'):
            continue
        user_id = int(prediction['user_id'])
        candidates_by_user.setdefault(user_id, []).append(
            (face_index, prediction, spoof_result)
        )

    conflicted_user_ids = {
        user_id for user_id, items in candidates_by_user.items() if len(items) > 1
    }
    for user_id in conflicted_user_ids:
        for _, prediction, _ in candidates_by_user[user_id]:
            _reset_consecutive_tracker(
                tracker_key, user_id=prediction.get('track_id', prediction.get('user_id'))
            )
    singletons_by_user = {
        user_id: items[0]
        for user_id, items in candidates_by_user.items()
        if len(items) == 1
    }
    counts = _sync_face_trackers(
        tracker_key,
        {
            item[1].get('track_id', item[1].get('user_id')): {
                'bbox': item[1].get('bbox'),
                'user_id': item[1].get('user_id'),
            }
            for item in singletons_by_user.values()
            if item[1].get('track_id', item[1].get('user_id')) is not None
        }
    )
    now = now_wib()
    hari = _get_nama_hari()
    waktu_sekarang = now.strftime('%H:%M:%S')
    tanggal_hari_ini = now.date()
    needs_schedule = any(
        count >= RECOGNITION_REQUIRED_FRAMES for count in counts.values()
    )
    try:
        jadwal_list = (
            _db_call_strict(db.get_jadwal_aktif, hari, waktu_sekarang)
            if needs_schedule else []
        )
    except db.DatabaseQueryError:
        jadwal_list = None

    face_results = []
    for face_index, prediction, spoof_result in prepared:
        spoof_result = {'is_real': True, 'label': 'DEFERRED', 'score': None}
        if prediction.get('recognition_status') == 'low_quality':
            response = {
                'status': 'error', 'tipe': 'low_quality',
                'pesan': 'Chat luong khuon mat chua dat; hay dieu chinh theo huong dan tren khung.',
                'quality_reason': prediction.get('quality_reason'),
            }
        elif not spoof_result.get('is_real'):
            tipe = 'no_face' if spoof_result.get('label') == 'NO_FACE' else 'spoofing'
            if tipe == 'spoofing':
                snapshot = _simpan_snapshot(frame, f'spoofing_{face_index}')
                db.catat_spoofing(snapshot, spoof_result.get('score', 0.0))
            response = {
                'status': 'skip' if tipe == 'no_face' else 'error',
                'tipe': tipe,
                'score': spoof_result.get('score', 0.0),
                'pesan': (
                    'KhuÃ´n máº·t khÃ´ng náº±m Ä‘á»§ trong vÃ¹ng áº£nh Ä‘á»ƒ kiá»ƒm tra.' if tipe == 'no_face'
                    else 'PhÃ¡t hiá»‡n hÃ nh vi giáº£ máº¡o! Vui lÃ²ng sá»­ dá»¥ng khuÃ´n máº·t tháº­t.'
                ),
                'spoofing': spoof_result
            }
        elif prediction.get('recognition_status') == 'needs_calibration':
            response = {
                'status': 'skip',
                'tipe': 'needs_calibration',
                'match_score': prediction.get('match_score'),
                'pesan': 'Cáº§n hiá»‡u chuáº©n ngÆ°á»¡ng cosine similarity trÆ°á»›c khi tá»± Ä‘á»™ng Ä‘iá»ƒm danh.',
                'spoofing': spoof_result
            }
        elif not prediction.get('dikenali'):
            response = {
                'status': 'error',
                'tipe': 'unknown',
                'confidence': prediction.get('confidence'),
                'match_score': prediction.get('match_score'),
                'pesan': f'KhÃ´ng nháº­n diá»‡n Ä‘Æ°á»£c khuÃ´n máº·t (Ä‘á»™ tin cáº­y: {prediction.get("confidence")}).',
                'spoofing': spoof_result
            }
        elif int(prediction['user_id']) in conflicted_user_ids:
            response = {
                'status': 'error',
                'tipe': 'identity_conflict',
                'confidence': prediction.get('confidence'),
                'pesan': 'Nhiá»u khuÃ´n máº·t bá»‹ gÃ¡n cÃ¹ng má»™t mÃ£ sinh viÃªn; khÃ´ng ghi Ä‘iá»ƒm danh cho danh tÃ­nh nÃ y.',
                'spoofing': spoof_result
            }
        else:
            user_id = int(prediction['user_id'])
            track_id = prediction.get('track_id', prediction.get('user_id'))
            cached_response = _get_completed_track_response(
                tracker_key, track_id, user_id
            )
            if cached_response:
                response = cached_response
                face_results.append(
                    _attach_face_metadata(response, prediction, face_index)
                )
                continue
            verification_count = counts.get(int(track_id), 0)
            if verification_count < RECOGNITION_REQUIRED_FRAMES:
                response = {
                    'status': 'skip',
                    'tipe': 'verifying',
                    'pesan': (
                        f'Äang xÃ¡c minh khuÃ´n máº·t... '
                        f'({verification_count}/{RECOGNITION_REQUIRED_FRAMES})'
                    ),
                    'verification_count': verification_count,
                    'required_frames': RECOGNITION_REQUIRED_FRAMES,
                    'spoofing': spoof_result
                }
            else:
                spoof_result = check_face(frame, prediction.get('bbox'))
                if not spoof_result.get('is_real'):
                    snapshot = _simpan_snapshot(frame, f'spoofing_{face_index}')
                    db.catat_spoofing(snapshot, spoof_result.get('score', 0.0))
                    # Consume the confirmation after a liveness failure so a
                    # replay cannot generate a spoof log on every frame.
                    _reset_consecutive_tracker(tracker_key, user_id=track_id)
                    response = {
                        'status': 'error', 'tipe': 'spoofing',
                        'score': spoof_result.get('score', 0.0),
                        'pesan': 'PhÃ¡t hiá»‡n hÃ nh vi giáº£ máº¡o! Vui lÃ²ng sá»­ dá»¥ng khuÃ´n máº·t tháº­t.',
                        'spoofing': spoof_result,
                    }
                else:
                    _reset_consecutive_tracker(tracker_key, user_id=track_id)
                    if jadwal_list is None:
                        response = _database_unavailable_response(spoof_result)
                    else:
                        try:
                            response = _process_verified_prediction(
                                frame, prediction, spoof_result, hari, waktu_sekarang,
                                tanggal_hari_ini, jadwal_list
                            )
                        except db.DatabaseQueryError:
                            response = _database_unavailable_response(spoof_result)
                    if (
                        response.get('data') and
                        (response.get('status') == 'ok' or response.get('tipe') == 'duplikat')
                    ):
                        _mark_completed_track(
                            tracker_key, track_id, user_id,
                            response['data'].get('jadwal_id'),
                            response
                        )
        face_results.append(
            _attach_face_metadata(response, prediction, face_index)
        )

    return _aggregate_face_results(face_results, tanggal_hari_ini)

@app.route('/api/face/health')
@login_required
def api_face_health():
    """Report model/gallery readiness without triggering model downloads."""
    from face.recognition import get_engine_health
    payload = get_engine_health()
    return jsonify(payload), 200 if payload.get('ready') else 503


@app.route('/api/absensi/proses', methods=['POST'])
@login_required
def api_absensi_proses():
    """Proses frame dari kamera untuk face recognition + absensi.

    Menerima base64 frame, jalankan anti-spoofing â†’ recognition â†’ catat absensi.
    """
    data = request.get_json()
    if not data or 'frame' not in data:
        return jsonify({'status': 'error', 'pesan': 'KhÃ´ng tÃ¬m tháº¥y khung hÃ¬nh.', 'data': None}), 400

    frame = _decode_frame(data['frame'])
    if frame is None:
        return jsonify({'status': 'error', 'pesan': 'KhÃ´ng thá»ƒ giáº£i mÃ£ khung hÃ¬nh.', 'data': None}), 400

    client_id = str(data.get('client_id', '')).strip()
    if not _valid_camera_client_id(client_id):
        return jsonify({
            'status': 'error', 'pesan': 'ID á»©ng dá»¥ng camera khÃ´ng há»£p lá»‡.', 'data': None
        }), 400
    tracker_key = _scanner_tracker_key(client_id)
    try:
        hasil = _proses_recognition(frame, tracker_key=tracker_key)
    except Exception as exc:
        from face.recognition import FaceEngineError
        if not isinstance(exc, FaceEngineError):
            raise
        return jsonify({
            'status': 'error',
            'tipe': 'model_unavailable',
            'pesan': str(exc),
            'results': []
        }), 503
    _broadcast_absensi_updates(hasil)
    return jsonify(hasil)


@app.route('/api/camera/toggle', methods=['POST'])
@login_required
def api_camera_toggle():
    """Toggle status kamera ON/OFF."""
    data = request.get_json(silent=True)
    if not isinstance(data, dict) or not isinstance(data.get('active'), bool):
        return jsonify({
            'status': 'error', 'data': None,
            'pesan': 'Tráº¡ng thÃ¡i mÃ¡y áº£nh pháº£i lÃ  giÃ¡ trá»‹ boolean.'
        }), 400

    active = data['active']
    client_id = str(data.get('client_id', '')).strip()
    state_key = _scanner_tracker_key(client_id) if client_id else f'http:{session.get("admin_id")}:default'
    _camera_states[state_key] = active
    if not active and _valid_camera_client_id(client_id):
        _reset_consecutive_tracker(_scanner_tracker_key(client_id), remove=True)
    return jsonify({
        'status': 'ok',
        'pesan': f'MÃ¡y áº£nh Ä‘Ã£ {"báº­t" if active else "táº¯t"}.',
        'data': {'camera_active': active}
    })


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# SERVE SNAPSHOT â€” Menyajikan foto bukti absensi
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@app.route('/snapshots/<path:filename>')
@login_required
def serve_snapshot(filename):
    """Sajikan file gambar snapshot bukti absensi."""
    from flask import send_from_directory
    return send_from_directory(SNAPSHOT_PATH, filename)


@app.route('/mahasiswa/<int:user_id>/foto')
@login_required
def serve_student_photo(user_id):
    """Sajikan satu thumbnail dataset tanpa membuka seluruh folder dataset."""
    from flask import abort, send_from_directory
    folder = os.path.join(DATASET_PATH, str(user_id))
    if not os.path.isdir(folder):
        abort(404)
    photos = sorted(f for f in os.listdir(folder) if f.lower().endswith('.jpg'))
    if not photos:
        abort(404)
    return send_from_directory(folder, photos[0])


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# WEBSOCKET HANDLERS (Flask-SocketIO)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@socketio.on('connect')
def handle_connect():
    """Client terhubung via WebSocket."""
    if 'admin_id' not in session:
        return False
    print('[SOCKET] Client terhubung.')


@socketio.on('disconnect')
def handle_disconnect():
    """Client terputus."""
    tracker_key = _socket_tracker_keys.pop(request.sid, None)
    if tracker_key:
        _reset_consecutive_tracker(tracker_key, remove=True)
    _camera_states.pop(f'socket:{request.sid}', None)
    print('[SOCKET] Client terputus.')


@socketio.on('camera_toggle')
def handle_camera_toggle(data):
    """Toggle kamera dari client."""
    if 'admin_id' not in session:
        emit('camera_status', {'active': False, 'error': 'unauthorized'})
        return
    active = data.get('active') if isinstance(data, dict) else None
    if not isinstance(active, bool):
        emit('camera_status', {'active': False, 'error': 'invalid_state'})
        return
    client_id = str(data.get('client_id', '')).strip()
    if client_id and not _valid_camera_client_id(client_id):
        emit('camera_status', {'active': False, 'error': 'invalid_client_id'})
        return
    tracker_key = _scanner_tracker_key(client_id) if client_id else request.sid
    _socket_tracker_keys[request.sid] = tracker_key
    _camera_states[f'socket:{request.sid}'] = active
    if not active:
        _reset_consecutive_tracker(tracker_key, remove=True)
        _socket_tracker_keys.pop(request.sid, None)
    emit('camera_status', {'active': active})


@socketio.on('process_frame')
def handle_process_frame(data):
    """Terima frame dari client via WebSocket, proses recognition."""
    if 'admin_id' not in session:
        emit('recognition_result', {
            'status': 'error',
            'pesan': 'Phiên đăng nhập đã hết hạn. Vui lòng đăng nhập lại.',
            'results': []
        })
        return

    try:
        if 'frame' not in data:
            emit('recognition_result', {'status': 'error', 'pesan': 'Khung hÃ¬nh trá»‘ng.'})
            return

        frame = _decode_frame(data['frame'])
        if frame is None:
            emit('recognition_result', {'status': 'error', 'pesan': 'KhÃ´ng thá»ƒ giáº£i mÃ£ khung hÃ¬nh.'})
            return

        client_id = str(data.get('client_id', '')).strip()
        if not _valid_camera_client_id(client_id):
            emit('recognition_result', {
                'status': 'error', 'tipe': 'invalid_client_id',
                'pesan': 'ID á»©ng dá»¥ng camera khÃ´ng há»£p lá»‡.', 'results': []
            })
            return
        tracker_key = _scanner_tracker_key(client_id)
        _socket_tracker_keys[request.sid] = tracker_key

        print(f'[SOCKET] process_frame: frame shape={frame.shape}')
        hasil = _proses_recognition(frame, tracker_key=tracker_key)
        print(f'[SOCKET] process_frame result: status={hasil.get("status")}, tipe={hasil.get("tipe", "-")}')
        emit('recognition_result', hasil)

        _broadcast_absensi_updates(hasil, skip_sid=request.sid)
    except Exception as e:
        from face.recognition import FaceEngineError
        if isinstance(e, FaceEngineError):
            emit('recognition_result', {
                'status': 'error', 'tipe': 'model_unavailable',
                'pesan': str(e), 'results': []
            })
            return
        print(f'[SOCKET ERROR] process_frame exception: {e}')
        import traceback
        traceback.print_exc()
        emit('recognition_result', {'status': 'error', 'pesan': f'Lá»—i mÃ¡y chá»§: {str(e)}'})



# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# AUTO-ALPHA: Background thread untuk tandai alpha otomatis
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _auto_alpha_checker():
    """Background thread: cek jadwal yang sudah selesai, tandai alpha otomatis."""
    import time as _time
    print('[AUTO-ALPHA] Background checker dimulai.')

    while True:
        _time.sleep(60)  # Cek setiap 60 detik
        try:
            hari = _get_nama_hari()
            waktu_sekarang = now_wib().strftime('%H:%M:%S')
            tanggal = now_wib().date()

            # Ambil semua jadwal yang sudah selesai hari ini
            jadwal_selesai = db.get_jadwal_selesai_hari_ini(hari, waktu_sekarang)

            for j in jadwal_selesai:
                jadwal_id = j['id']
                kelas_id = j['kelas_id']

                # Cari mahasiswa yang belum absen
                belum_absen = db.get_mahasiswa_belum_absen(jadwal_id, kelas_id, tanggal)

                if belum_absen:
                    user_ids = [m['id'] for m in belum_absen]
                    count = db.bulk_catat_alpha(jadwal_id, user_ids, tanggal)
                    if count > 0:
                        nama_mk = j.get('nama_mk', '?')
                        print(f'[AUTO-ALPHA] {count} mahasiswa ditandai alpha '
                              f'untuk {nama_mk} (jadwal #{jadwal_id})')

        except Exception as e:
            print(f'[AUTO-ALPHA] Error: {e}')


def _start_background_tasks_once():
    """Start background jobs once per process, termasuk saat dimuat WSGI."""
    global _background_started
    with _background_lock:
        if _background_started:
            return
        alpha_thread = threading.Thread(target=_auto_alpha_checker, daemon=True)
        alpha_thread.start()
        _background_started = True


# Gunicorn mengimpor modul tanpa menjalankan blok __main__. Mulai job saat
# worker dibuat; pada debug reloader hanya proses anak yang menjalankannya.
if not FLASK_DEBUG or os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
    _start_background_tasks_once()


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# ENTRY POINT
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•


if __name__ == '__main__':
    # Jalankan auto-alpha checker di background thread
    _start_background_tasks_once()

    print("=" * 50)
    print("   FLASK + SOCKETIO â€” SISTEM ABSENSI")
    print("=" * 50)
    print(f"\n[INFO] Dashboard  : http://127.0.0.1:{FLASK_PORT}")
    print(f"[INFO] Login      : http://127.0.0.1:{FLASK_PORT}/login")
    print(f"[INFO] WebSocket  : ws://127.0.0.1:{FLASK_PORT}")
    print(f"[INFO] Anti-Spoof : threshold={ANTI_SPOOFING_THRESHOLD}")
    print(f"[INFO] Confidence : threshold={CONFIDENCE_THRESHOLD}")
    print(f"[INFO] ESP32      : {'Aktif' if ESP32_ENABLED else 'Nonaktif'}")
    print(f"[INFO] Auto-Alpha : Aktif (cek setiap 60 detik)")
    print(f"[INFO] Tekan Ctrl+C untuk menghentikan.\n")
    socketio.run(app, host=FLASK_HOST, port=FLASK_PORT, debug=FLASK_DEBUG,
                 allow_unsafe_werkzeug=True)
